"""
MNR Template Agent — port of MNR_Setup_vRG_KeplerCannon_Apr2026.py.
==================================================================

End-to-end: Contract PDF → forensic fee extraction → material-code matching
→ SAP-ready MNR Excel draft with biller colour-coding.

Pipeline (mirrors the original script line for line):

    STAGE 1  Forensic per-page extraction (600 DPI, checkbox-aware,
             12-page chunks). Captures: item, price (numeric or status
             label), quantity, frequency, pricing_condition, section_header,
             page, checkbox_checked, explanation.

    STAGE 2  Match each extracted item to a Portico material code.
             Primary dictionary (Frequently Used Material Codes) is
             preferred; the full Portico master is the fallback when no
             primary candidate reaches confidence >= 0.70.

    STAGE 3  Build the SAP MNR draft (one row per fee line) with the
             biller-known fields BLANK and rows colour-coded by match
             provenance (green / blue / orange / pink / grey, with a
             dark-orange MATERIAL overlay for PMIN-minimum or known
             placeholder codes).

Difference from the standalone script (the only behavioural change):

    Caches are namespaced PER CLIENT under
        Output/<Client>/.mnr_cache/<pdf_stem>.stage1.json
        Output/<Client>/.mnr_cache/<pdf_stem>.stage2.json
    The standalone script kept caches next to itself, which forced the
    user to delete them before processing a different client. Per-client
    paths fix that — runs for different clients can never collide.

Agent contract (matches the other 7 frontend agents):
    run(client_name, api_key="", model="", progress_callback=None,
        contracts=None, core="", dictionary_path=None,
        force=False, **_legacy_kwargs) -> dict
    is_processed(client_name) -> bool
    output_path(client_name)  -> Path
"""

from __future__ import annotations

import base64
import calendar
import io
import json
import math
import os
import re
import shutil
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Callable, Optional

import fitz                                       # PyMuPDF
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image

from fiserv_client import make_client

# Core paths / helpers — these have existed in config.py for the lifetime of
# the project, so a hard import is safe.
from config import (
    BASE_DIR,
    INPUT_DIR,
    OUTPUT_DIR,
    client_input_dir,
    default_dictionary_for,
)

# MNR-specific constants — defensive imports with env-var fallbacks. This
# matters because a stale `__pycache__/config.cpython-XYZ.pyc` (compiled
# before the MNR keys existed) can otherwise raise ImportError at startup
# even though the live config.py defines every key. With these fallbacks the
# agent runs from env vars or sensible defaults whenever any single constant
# fails to import — laptop (OpenAI) and Fiserv VDI alike.
try:
    from config import MNR_API_KEY                                # type: ignore
except ImportError:
    # Laptop default: piggyback on the chatbot-wide extraction key.
    # Fiserv VDI: the Foundation endpoint authenticates by VDI network +
    # X-Email-Id header instead, so this value is unused there.
    try:
        from config import EXTRACTION_API_KEY as MNR_API_KEY      # type: ignore
    except ImportError:
        MNR_API_KEY = os.environ.get("MNR_API_KEY", "")

try:
    from config import MNR_MODEL                                  # type: ignore
except ImportError:
    MNR_MODEL = os.environ.get("MNR_MODEL", "gpt-5.2-2025-12-11")

try:
    from config import MNR_FREQ_CATALOG_NAME                      # type: ignore
except ImportError:
    MNR_FREQ_CATALOG_NAME = os.environ.get(
        "MNR_FREQ_CATALOG_NAME", "Frequently Used Material Codes.xlsx"
    )

try:
    from config import MNR_CHECKBOX_REF                           # type: ignore
except ImportError:
    MNR_CHECKBOX_REF = BASE_DIR / "marked_checkbox_example.png"

try:
    from config import MNR_TEMPLATE_NAME                          # type: ignore
except ImportError:
    MNR_TEMPLATE_NAME = os.environ.get("MNR_TEMPLATE_NAME", "MNR_template.xlsx")

try:
    from config import MNR_DPI                                    # type: ignore
except ImportError:
    MNR_DPI = int(os.environ.get("MNR_DPI", "600"))

try:
    from config import MNR_CHUNK_SIZE                             # type: ignore
except ImportError:
    MNR_CHUNK_SIZE = int(os.environ.get("MNR_CHUNK_SIZE", "12"))

try:
    from config import MNR_MATCH_MIN_CONF                         # type: ignore
except ImportError:
    MNR_MATCH_MIN_CONF = float(os.environ.get("MNR_MATCH_MIN_CONF", "0.70"))


_VERSION = "mnr_template/1.0 (port of MNR_Setup_vRG_KeplerCannon_Apr2026.py)"


# ─────────────────────────────────────────────────────────────────────────────
# Public path helpers (used by chatbot.py for the agent registry).
# ─────────────────────────────────────────────────────────────────────────────

def output_path(client_name: str) -> Path:
    """Per-client SAP-ready MNR Excel draft."""
    return OUTPUT_DIR / client_name / "mnr_output.xlsx"


def is_processed(client_name: str) -> bool:
    return output_path(client_name).exists()


def _cache_dir(client_name: str) -> Path:
    """Per-client cache root. Created on demand. This is the directory that
    used to be 'BASE next to the .py script' in the standalone version."""
    d = OUTPUT_DIR / client_name / ".mnr_cache"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _stage1_cache(client_name: str, pdf_stem: str) -> Path:
    return _cache_dir(client_name) / f"{pdf_stem}.stage1.json"


def _stage2_cache(client_name: str, pdf_stem: str) -> Path:
    return _cache_dir(client_name) / f"{pdf_stem}.stage2.json"


def _dropped_path(client_name: str, pdf_stem: str) -> Path:
    return _cache_dir(client_name) / f"{pdf_stem}.dropped_legal.json"


def clear_caches(client_name: str) -> int:
    """Wipe the per-client cache directory. Returns the number of files
    removed. Called by the agent itself when `force=True` is passed — the
    user never has to do this by hand."""
    d = OUTPUT_DIR / client_name / ".mnr_cache"
    if not d.exists():
        return 0
    n = sum(1 for _ in d.glob("*"))
    shutil.rmtree(d, ignore_errors=True)
    return n


# ─────────────────────────────────────────────────────────────────────────────
# Catalog / dictionary discovery.
# ─────────────────────────────────────────────────────────────────────────────

def _find_freq_catalog(core: str) -> Optional[Path]:
    """Locate the Frequently Used Material Codes dictionary.
    1) Input/<Core>/<MNR_FREQ_CATALOG_NAME>
    2) BASE_DIR / <MNR_FREQ_CATALOG_NAME>
    """
    if core:
        p = INPUT_DIR / core / MNR_FREQ_CATALOG_NAME
        if p.exists():
            return p
    p = BASE_DIR / MNR_FREQ_CATALOG_NAME
    return p if p.exists() else None


def _find_template(client_name: str, core: str) -> Optional[Path]:
    """Find an MNR header template, in this order:
    1) Input/<Core>/<Client>/<MNR_TEMPLATE_NAME>
    2) Input/<Core>/<MNR_TEMPLATE_NAME>
    3) BASE_DIR / <MNR_TEMPLATE_NAME>
    """
    candidates = []
    if core:
        candidates += [
            client_input_dir(client_name, core) / MNR_TEMPLATE_NAME,
            INPUT_DIR / core / MNR_TEMPLATE_NAME,
        ]
    candidates += [BASE_DIR / MNR_TEMPLATE_NAME]
    return next((p for p in candidates if p.exists()), None)


# Default column header set used when no template is found. Matches the SAP
# MNR draft columns that the script's build_mnr_rows() writes.
_DEFAULT_TEMPLATE_COLUMNS = [
    "MATERIAL", "Desc", "Qty", "Item Cat", "Start Date", "End Date",
    "Cond.Type", "Price", "Currency", "Cust. Mat. Num.", "Mat.Grp.2",
    "Contract signed", "Bill Plan Date",
]


# ─────────────────────────────────────────────────────────────────────────────
# Signing date — from filename, exactly like the script.
# ─────────────────────────────────────────────────────────────────────────────

def _signing_date_from_filename(path: str) -> Optional[date]:
    """Parse a M-D-YYYY-style date out of a contract filename.

    Unlike the original script (which hard-rejected pre-2022 dates because
    it was tailored to one biller's recent-contracts workflow), this helper
    accepts any plausible year (1990-2100) so the MNR agent can pick the
    latest master agreement on file even for clients whose master was
    signed long ago. Year filtering is the chatbot's responsibility, not
    this date parser's."""
    m = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', os.path.basename(path))
    if not m:
        return None
    mo, d, y = [int(x) for x in m.groups()]
    if not (1990 <= y <= 2100):
        return None
    try:
        return date(y, mo, d)
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# CPI Eligible column maintenance on the Frequently Used dictionary.
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_cpi_column(catalog_path: Path, log: Callable[[str], None]) -> None:
    """Add a 'CPI Eligible' column to the Frequently Used dictionary if it
    isn't already present. Verbatim from the script's ensure_cpi_column()."""
    try:
        existing_headers = list(pd.read_excel(catalog_path, nrows=0).columns)
        if "CPI Eligible" in existing_headers:
            return
    except Exception:
        pass

    try:
        wb = load_workbook(str(catalog_path))
    except PermissionError:
        log("WARNING: Frequently Used dictionary is locked by another "
            "process; skipping CPI-column setup.")
        return
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if "CPI Eligible" in headers:
        wb.close()
        return
    new_col = len(headers) + 1
    ws.cell(row=1, column=new_col, value="CPI Eligible")

    def col_idx(name):
        try:    return headers.index(name) + 1
        except: return None
    desc_c = col_idx("Description")
    ic_c   = col_idx("Item Category")

    def guess_cpi(desc, item_cat):
        d = (desc or "").upper()
        ic = (item_cat or "").upper()
        if ic == "ZINL":
            return "No"
        for kw in ("POSTAGE", "TRAVEL", "PASS THROUGH", "PASS-THROUGH",
                   "THIRD PARTY", "THIRD-PARTY", "VENDOR"):
            if kw in d:
                return "No"
        return "Yes"

    for row in range(2, ws.max_row + 1):
        desc = ws.cell(row=row, column=desc_c).value if desc_c else ""
        ic   = ws.cell(row=row, column=ic_c).value   if ic_c   else ""
        ws.cell(row=row, column=new_col, value=guess_cpi(desc, ic))

    wb.save(str(catalog_path))
    wb.close()
    log("Added 'CPI Eligible' column to Frequently Used dictionary.")


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — forensic extraction
# ─────────────────────────────────────────────────────────────────────────────

STAGE1_SYSTEM_PROMPT = """
You are a forensic-level contract analysis expert.

Your task is to extract ALL fee line items from the contract. This includes:
  (A) items with an explicit dollar amount (e.g. "$2,750.00", "$0.15 per page"), AND
  (B) items whose price cell shows a NON-DOLLAR STATUS LABEL instead of a number.
      Capture these too - the biller decides whether to bill them later.
      Typical non-dollar statuses to capture AS-WRITTEN in the `price` field:
          "Included", "Bundled", "Bundled Fee", "TBD", "Prev Paid",
          "Waived", "N/A", "No Charge", "Complimentary", "Free", "$0.00"
      Do NOT invent a number for these - keep the literal status string.

Still SKIP:
  - strikethrough values with no replacement
  - strikethrough values labeled Included/Waived/N/A (truly non-billable)
  - contextual / threshold dollar figures (asset size, etc. - see EXCLUSIONS below)
  - totals/subtotals

CRITICAL INSTRUCTIONS FOR CHECKBOX DETECTION:

1. CHECKBOX PLACEMENT: Checkboxes can appear BEFORE or AFTER the fee description.
2. CHECKBOX TYPES:
   - CHECKED: filled box, box with any mark inside
   - UNCHECKED: empty square outline, no marks inside
3. HIERARCHICAL STRUCTURE: parent checkboxes govern child items.

ADVANCED CHECKBOX ASSOCIATION: look upward within the same visual block for the
nearest preceding checkbox. Parent/section checkboxes propagate to child items
and table rows. Row-level checkboxes override table-level ones. Section-level
checkboxes apply until a new numbered section or major heading begins.

TABLE-LEVEL ASSOCIATION: a checkbox immediately above a table (no section break
between) governs the entire table unless a row has its own checkbox.

NUMBERED-SECTION INHERITANCE: a checkbox next to a section heading governs all
subordinate content (lettered/roman sub-clauses, indented fees, sub-tables).

SECTION HEADER IDENTIFICATION: for each extracted item, identify the nearest
section header (centered / bold / top of page, often multi-line). Inherit across
pages until a new header appears. Return the header text verbatim.

STRIKETHROUGH: ignore crossed-out dollar values. If a new price follows on the
same line, capture only the new price. If BOTH the number IS struck through
AND the text is labeled Included/Waived/N/A/Free, omit that item. (When the
label is Included/Bundled/TBD/Prev Paid without a strikethrough, KEEP the
item per rule (B) above - store the label verbatim in the `price` field.)

EXCLUSIONS:
  - Totals/Subtotals/Grand totals.
  - Non-fee figures (asset size, account balance, threshold triggers).
  - Conditional threshold labels ("Over $100M", "more than $X"); only extract
    the actual fee amount, not the trigger.
  - LEGAL / COMMITMENT / LIABILITY DOLLAR FIGURES. Do NOT extract any dollar
    amount that sits in a non-fee clause describing Fiserv's own liability,
    SLA credit, indemnification cap, late-payment commitment, or similar
    reimbursement obligation. Tell-tale phrases include:
        "Fiserv will bear", "Fiserv's responsibility", "Fiserv's liability",
        "Fiserv agrees to reimburse", "Fiserv will credit up to",
        "Late Payment Commitment", "SLA credit", "indemnif", "limitation of
        liability", "up to $X" (when $X is a liability cap rather than a
        billable fee), "Service Level Credits".
    These clauses describe what FISERV pays the client under fault, not what
    the CLIENT pays for a service. They are not fee items.

TABLE COLUMN PRIORITY:
  - QTY + UNIT PRICE + AMOUNT: use AMOUNT as price, QTY as quantity, units as
    frequency. Ignore UNIT PRICE. One row per line item.
  - If AMOUNT is TBD/blank: fall back to UNIT PRICE.
  - No AMOUNT column: use the sole price column.

QUANTITY: only explicit counts ("3", "5 users"). Tier bands / asset ranges /
"First 200 checks" go into pricing_condition, not quantity. Unit descriptors
like "Per FI"/"Per User" go into frequency, not quantity.

FREQUENCY: time- or usage-based cadence, preserved verbatim.

PRICING_CONDITION: installments, triggers, tier bands, thresholds, minimums.

OUTPUT: strict JSON, no prose, no markdown:

{
  "items": [
    {
      "item": "<clear description>",
      "checkbox_checked": true | false | null,
      "price": "<dollar value exactly as written>",
      "quantity": "<qty or null>",
      "frequency": "<frequency or null>",
      "pricing_condition": "<condition or null>",
      "page": <page number>,
      "explanation": "brief note on checkbox location and state",
      "section_header": "<section title or null>"
    }
  ]
}

RULES:
- Empty box = checkbox_checked: false; filled = true; no checkbox at all = null.
- Use only visible text.
- Do not infer missing values.
- Be exhaustive. Missing a dollar value is an error. Missing an Included/
  Bundled/TBD/Prev Paid/No Charge/Waived line is ALSO an error.
- Every visible fee-line row of a fee table or fee schedule must produce an
  item entry, whether it shows a number or a status label.
"""


def _pdf_to_images(pdf_path: str, dpi: int = MNR_DPI) -> list[dict]:
    doc = fitz.open(pdf_path)
    pages = []
    for i in range(len(doc)):
        page = doc.load_page(i)
        pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72), alpha=False)
        img = Image.open(BytesIO(pix.tobytes("png")))
        pages.append({"page_number": i + 1, "image": img})
    doc.close()
    return pages


def _image_to_b64(img) -> str:
    buf = BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def _extract_chunk(client, page_chunk, chunk_index, total_chunks,
                   checkbox_ref_path: Path, model_name: str) -> dict:
    content = []
    with open(checkbox_ref_path, "rb") as f:
        ref_b64 = base64.b64encode(f.read()).decode("utf-8")
    content.append({
        "type": "input_text",
        "text": "First image is a REFERENCE MARKED checkbox. A square with "
                "two crossing diagonal lines (X) is always MARKED, not empty."
    })
    content.append({
        "type": "input_image",
        "image_url": f"data:image/png;base64,{ref_b64}",
    })
    content.append({
        "type": "input_text",
        "text": f"The following images are pages from a contract, "
                f"chunk {chunk_index}/{total_chunks}."
    })
    for p in page_chunk:
        content.append({
            "type": "input_image",
            "image_url": f"data:image/png;base64,{_image_to_b64(p['image'])}",
        })
    r = client.responses.create(
        model=model_name,
        input=[
            {"role": "system", "content": STAGE1_SYSTEM_PROMPT},
            {"role": "user",   "content": content},
        ],
    )
    try:
        return json.loads(r.output_text)
    except Exception as e:
        m = re.search(r"\{[\s\S]*\}", r.output_text)
        if m:
            try:    return json.loads(m.group(0))
            except: pass
        return {"items": []}


def _stage1_extract(client, pdf_path: str, cache_path: Path,
                    checkbox_ref_path: Path, model_name: str,
                    log: Callable[[str], None]) -> list[dict]:
    """Run Stage 1, or load the cached result if `cache_path` exists."""
    if cache_path.exists():
        log(f"  Using cached Stage 1: {cache_path.name}")
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)

    pages = _pdf_to_images(pdf_path)
    total_pages = len(pages)
    total_chunks = math.ceil(total_pages / MNR_CHUNK_SIZE) or 1
    log(f"  Stage 1: {total_pages} pages in {total_chunks} chunks at {MNR_DPI} DPI.")

    items: list[dict] = []
    for i in range(total_chunks):
        chunk = pages[i * MNR_CHUNK_SIZE : (i + 1) * MNR_CHUNK_SIZE]
        log(f"    chunk {i+1}/{total_chunks} (pages "
            f"{chunk[0]['page_number']}-{chunk[-1]['page_number']})…")
        parsed = _extract_chunk(client, chunk, i + 1, total_chunks,
                                checkbox_ref_path, model_name)
        items.extend(parsed.get("items", []))

    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2)
    log(f"  Stage 1: extracted {len(items)} line items.")
    return items


# Belt-and-braces filter for legal/SLA/liability clause dollar amounts. Same
# regex as the script's drop_legal_clause_items().
_LEGAL_CLAUSE_RE = re.compile(
    r"(?:"
    r"fiserv\s+will\s+bear"
    r"|fiserv\s+will\s+credit"
    r"|fiserv\s+agrees\s+to\s+reimburse"
    r"|fiserv['’]?s\s+(?:responsibility|liability)"
    r"|late\s+payment\s+commitment"
    r"|service\s+level\s+credits?"
    r"|sla\s+credits?"
    r"|indemnif"
    r"|limitation\s+of\s+liability"
    r"|liability\s+cap"
    r"|aggregate\s+liability"
    r")",
    re.I,
)


def _drop_legal_clause_items(items: list[dict], dropped_path: Path,
                             log: Callable[[str], None]) -> list[dict]:
    kept, dropped = [], []
    for it in items:
        blob = " ".join(str(it.get(k) or "") for k in
                        ("item", "section_header", "pricing_condition",
                         "explanation"))
        if _LEGAL_CLAUSE_RE.search(blob):
            dropped.append(it)
        else:
            kept.append(it)
    if dropped:
        log(f"  Filtered {len(dropped)} legal-clause items "
            f"(see {dropped_path.name}).")
        with open(dropped_path, "w", encoding="utf-8") as f:
            json.dump(dropped, f, indent=2)
    return kept


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — material code matching
# ─────────────────────────────────────────────────────────────────────────────

STAGE2_SYSTEM_PROMPT_TEMPLATE = """
You are a material-code matching assistant for Fiserv Portico MNR setup.

You receive:
  - EXTRACTED ITEMS from a contract (section_header, item text, frequency,
    pricing_condition, page).
  - PRIMARY dictionary (Frequently Used Material Codes), GROUPED BY Material
    Code. Many codes have multiple allowed descriptions - pick the one that
    best matches the contract's section_header + item text.
  - FALLBACK dictionary (Portico master): flat list of Material Code +
    Description.

FOR EACH extracted item, pick the SINGLE best-matching material code:
  * ALWAYS try to match. Match every item you can. Do NOT return null just
    because the item belongs to a section another team might bill - the
    reviewer will filter. Only return null when truly no dictionary entry
    is a reasonable semantic match.
  * Prefer a PRIMARY-dictionary match over a fallback match.
  * Only use the FALLBACK dictionary when NO primary candidate is a
    reasonable match (confidence >= {match_min_conf}).
  * If you still can't find any reasonable match, return material_code = null.

DISAMBIGUATION for codes with multiple descriptions (e.g. CUPR1415 covers
SSL, telecom, network, 4G…): use the item's section_header + item text to
pick the correct description. Return that specific description in `description`.

Multi-line table cells: if a cell contains a product name header followed by
a sub-descriptor, match to the PRODUCT NAME. Example:
   section='Digital Online Banking Services'
   item='Virtual Branch Next / Scheduled transfer conversion'
   price='$3,600 One Time'
   -> VBNCUPR0002 'VIRTUAL BRANCH NEXT IMPLEMENTATION' (NOT VBNCUPR0003).

IMPORTANT: `pricing_condition` is EITHER:
  (a) a WHEN clause describing a tier/threshold/timing:
      e.g. 'First 200 checks', 'Asset size up to $100M',
           'Installments: 50% upfront, 50% on completion',
           '201 to 1,000 checks deposited'
  (b) a WHAT clause - additional product description that extraction split off
      from the main item text:
      e.g. 'Core migration for existing Virtual Branch clients only',
           'Bill Discovery Services - Implementation Fee',
           'Used for CheckFree Hosted UI core migration'

When pricing_condition is type (b), treat it as part of the product name and
match on the COMBINED text. Stage-1 extraction often misaligns multi-line
table cells so the actual product label ends up in pricing_condition while a
sub-descriptor (like 'Scheduled transfer conversion') ends up in item.

Example (typical fix):
   item='Scheduled transfer conversion'
   pricing_condition='Core migration for existing Virtual Branch clients only'
   -> Combined product = 'Core migration for existing Virtual Branch clients'
   -> VBNCUPR0005 'CORE MIGRATION EXISTING VBN OLB CLIENTS' (NOT CUPR0985
      'VBN SCHEDULED TRANSFERS MERGER').

KEYWORD-OVERLAP RULE: prefer codes whose description contains the SAME
qualifying tokens from the contract item text. Two kinds of tokens matter:

  (1) PRODUCT-FAMILY tokens (identify WHAT product the line is about).
      These families are DISTINCT - do NOT cross between them:
      WISDOM, LOANCIERGE, MOBILITI, VBN / VIRTUAL BRANCH, PORTICO, NAUTILUS,
      CHECKFREE, FILE EXCHANGE, FOS (FISERV OUTPUT SOLUTIONS), FRAUDNET,
      SECURENOW, CONVERGEIT / IVR, STATEMENT EXPRESS, DOCUMENT IT,
      RELATIONSHIP PRICING, CTR, ANB (NETWORK BALANCING), RA (REPORTING
      ANALYTICS), FUND NOW, DEPOSIT ESSENTIALS / ORIGINATE, WEB SIGNATURE,
      ENHANCED SKIP PAY, ID CAPTURE, CREDIT BUREAU,
      ID VERIFICATION / IDV   (CUPR0673 monthly, CUPR0674 per-txn)  <- DISTINCT
      CONFIRM IT / OFAC / FINCEN  (CUPR0675)                         <- DISTINCT
        ^ these last two are SEPARATE products; do not substitute one
          for the other just because both involve identity/verification.

  (2) FEE-TYPE tokens (identify WHAT KIND of fee within the family):
      - APP / APPLICATION      -> the one-time app fee
      - IMPL / IMPLEMENTATION  -> setup / implementation fee
      - SETUP / SET-UP / SET UP -> setup fee
      - OT / ONE TIME / ONE-TIME -> one-time variant
      - MONTHLY / MAINT / MAINTENANCE / HOSTING -> ongoing recurring fee
      - TRANSACTION / PER-TXN / USER DEVICE -> per-usage fee
      - SSO / LINKS / LINK     -> SSO integration, NOT a migration
      - HOSTED                 -> hosted variant
      - BUNDLE / BUNDLED       -> bundled pricing
      - MERGER                 -> core-system merger, NOT SSO link
      - MIGRATION / CONVERSION -> data/core migration
      - RENEWAL                -> renewal of existing service
      - TIER / ASSET SIZE      -> tier-based pricing

STRICT RULES:

  (A) The matched code's description MUST share the PRODUCT-FAMILY token with
      the contract line's item+section_header. If the contract mentions
      'SecureNow' and no primary code has SECURENOW in its description, DO NOT
      pick a code from a different family (e.g. CONVERGEIT / IVR). Return
      material_code = null instead.

  (B) Within the right family, the matched code's description MUST share the
      FEE-TYPE token with the contract line. If the contract says
      'Application Fee' pick the APP/ONE-TIME variant, NOT the MONTHLY/HOSTING
      variant. Example:
         item     = 'Loancierge Application Fee'
         section  = 'Consumer Loan Origination'
         price    = $7,200 (status Included)
      -> CUPR0828 'LOANCIERGE APP OT FEE'      (has LOANCIERGE + APP + OT)
      NOT CUPR0829 'LOANCIERGE MONTHLY HOSTING FEE' (wrong fee-type).

  (C) NO-MATCH IS BETTER THAN WRONG-FAMILY. If no code in either dictionary
      shares BOTH the product-family token AND the fee-type token, return
      material_code = null. Do not reach across families just to produce
      a code - the biller will fill in a correct code or skip the row.

  (D) Exception: 'Portico' fees often appear WITHOUT the word 'Portico' in the
      item text (they live under 'Account Processing Services (Portico)
      Schedule' section headers). Use the section_header to identify the
      Portico family in that case.

OUTPUT: strict JSON, one object per extracted item, SAME ORDER and SAME COUNT
as the input. Do not drop items.
{{
  "matched": [
    {{
      "idx":            int,
      "material_code":  str | null,
      "description":    str | null,
      "item_category":  str | null,
      "source_dict":    "primary" | "fallback" | "none",
      "confidence":     float,
      "reasoning":      str
    }}
  ]
}}
"""


def _extract_minimum_fee(pricing_condition) -> Optional[float]:
    if not pricing_condition:
        return None
    m = re.search(r"minimum[^$]*\$?\s*([\d,]+\.?\d*)",
                  str(pricing_condition), re.I)
    if m:
        try:    return float(m.group(1).replace(",", ""))
        except: return None
    return None


_NON_NUMERIC_STATUSES = {
    "included", "bundled", "bundled fee", "tbd", "prev paid", "previously paid",
    "waived", "waived fee", "n/a", "na", "no charge", "nocharge",
    "free", "complimentary", "pass through", "pass-through",
}


def _parse_price(val):
    if val is None: return None, ""
    if isinstance(val, (int, float)): return float(val), ""
    raw = str(val).strip()
    s = raw.replace("$", "").replace(",", "").replace("USD", "").strip()
    if not s: return None, ""
    low = s.lower()
    if low in _NON_NUMERIC_STATUSES:
        return None, raw.strip()
    try:
        return float(s), ""
    except Exception:
        return None, raw.strip()


def _classify_item_category(freq_text) -> str:
    f = (freq_text or "").lower()
    if any(k in f for k in ("one time", "one-time", "onetime")):
        return "ZINL"
    if any(k in f for k in ("monthly", "/month", "per month", "annual", "per year")):
        if any(k in f for k in ("per ", "/trans", "per item", "each")):
            return "ZINR"
        return "ZINM"
    if any(k in f for k in ("per transaction", "per item", "per user",
                            "per account", "per call", "per envelope",
                            "per device", "each")):
        return "ZINR"
    return "ZINR"


def _group_primary_dict(records: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for r in records:
        code = (r.get("Material Code") or "").strip()
        if not code:
            continue
        if code not in grouped:
            grouped[code] = {
                "material_code":  code,
                "descriptions":   [],
                "item_category":  r.get("Item Category")  or "",
                "condition_type": r.get("Condition Type") or "",
                "cpi_eligible":   r.get("CPI Eligible")   or "",
            }
        desc = (r.get("Description") or "").strip()
        if desc and desc not in grouped[code]["descriptions"]:
            grouped[code]["descriptions"].append(desc)
    return list(grouped.values())


def _stage2_match(client, items: list[dict], primary_records: list[dict],
                  fallback_records: list[dict], cache_path: Path,
                  model_name: str, log: Callable[[str], None]) -> list[dict]:
    if cache_path.exists():
        log(f"  Using cached Stage 2: {cache_path.name}")
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)

    compact = []
    for it in items:
        compact.append({
            "idx":               len(compact),
            "item":              it.get("item"),
            "section_header":    it.get("section_header"),
            "frequency":         it.get("frequency"),
            "pricing_condition": it.get("pricing_condition"),
            "page":              it.get("page"),
        })

    primary_grouped = _group_primary_dict(primary_records)

    BATCH = 25
    batches = [compact[i:i+BATCH] for i in range(0, len(compact), BATCH)]
    log(f"  Stage 2: matching {len(compact)} items in {len(batches)} "
        f"batch(es) of <= {BATCH}…")

    stage2_sys_prompt = STAGE2_SYSTEM_PROMPT_TEMPLATE.format(
        match_min_conf=MNR_MATCH_MIN_CONF
    )

    all_matched: list[dict] = []
    for bi, batch in enumerate(batches):
        user_payload = (
            "PRIMARY DICTIONARY (grouped - each code has one or more allowed descriptions):\n"
            f"{json.dumps(primary_grouped, ensure_ascii=False)}\n\n"
            "FALLBACK DICTIONARY (Portico master - flat):\n"
            f"{json.dumps(fallback_records, ensure_ascii=False)}\n\n"
            "EXTRACTED ITEMS FOR THIS BATCH (preserve order and idx in your output):\n"
            f"{json.dumps(batch, ensure_ascii=False)}"
        )
        log(f"    batch {bi+1}/{len(batches)} ({len(batch)} items)…")
        r = client.responses.create(
            model=model_name,
            input=[
                {"role": "system", "content": stage2_sys_prompt},
                {"role": "user",   "content": [
                    {"type": "input_text", "text": user_payload}
                ]},
            ],
        )
        try:
            out = json.loads(r.output_text)
        except Exception:
            m = re.search(r"\{[\s\S]*\}", r.output_text)
            out = json.loads(m.group(0)) if m else {"matched": []}
        batch_matches = out.get("matched", [])
        by_idx = {m.get("idx"): m for m in batch_matches if isinstance(m, dict)}
        for bitem in batch:
            bi_idx = bitem["idx"]
            if bi_idx in by_idx:
                all_matched.append(by_idx[bi_idx])
            else:
                all_matched.append({
                    "idx": bi_idx, "material_code": None,
                    "source_dict": "none", "confidence": 0.0,
                    "reasoning": "dropped by matcher",
                })

    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(all_matched, f, indent=2)
    log(f"  Stage 2: produced {len(all_matched)} match record(s).")
    return all_matched


# ─────────────────────────────────────────────────────────────────────────────
# Stage 3 — build MNR rows
# ─────────────────────────────────────────────────────────────────────────────

def _build_mnr_rows(items, matched, signing_date: date,
                    output_columns: list[str],
                    primary_records: list[dict],
                    master_lookup: dict) -> tuple[list[dict], list[dict], list[str]]:
    def first_of_next_month(d):
        if d.month == 12: return date(d.year + 1, 1, 1)
        return date(d.year, d.month + 1, 1)

    start_date = first_of_next_month(signing_date)
    end_date   = start_date + relativedelta(years=10)
    last_day_start = date(start_date.year, start_date.month,
                          calendar.monthrange(start_date.year, start_date.month)[1])

    signing_short = f"{signing_date.month}/{signing_date.day}/{signing_date.year}"
    cust_note_base = f"AGRMNT DTD {signing_short}"

    extra_cols = ["Checkbox", "Contract Page", "Price Status", "Source Contract"]
    full_cols  = list(output_columns) + extra_cols

    def blank_row():
        return {c: "" for c in full_cols}

    rows: list[dict] = []
    meta: list[dict] = []

    while len(matched) < len(items):
        matched.append({"material_code": None, "source_dict": "none",
                        "confidence": 0.0})

    for it, mt in zip(items, matched):
        price_val, price_status = _parse_price(it.get("price"))
        freq      = it.get("frequency") or ""
        cond      = it.get("pricing_condition") or ""
        section   = it.get("section_header") or ""
        desc_raw  = it.get("item") or ""
        page_n    = it.get("page")
        checkbox  = it.get("checkbox_checked")
        src_pdf   = it.get("_source_pdf") or ""

        mat_code    = (mt or {}).get("material_code")
        source_dict = (mt or {}).get("source_dict") or "none"
        model_ic    = (mt or {}).get("item_category")
        picked_desc = ((mt or {}).get("description") or "").strip()

        dict_desc = ""
        item_cat  = model_ic or _classify_item_category(freq)
        cond_type = "ZPRM"
        cpi_flag  = "Yes"

        if mat_code:
            prim_rows = [r for r in primary_records
                         if (r.get("Material Code") or "").strip() == mat_code]
            if prim_rows:
                chosen = None
                if picked_desc:
                    for r in prim_rows:
                        if (r.get("Description") or "").strip().upper() == picked_desc.upper():
                            chosen = r; break
                if chosen is None:
                    chosen = prim_rows[0]
                dict_desc = chosen.get("Description") or ""
                item_cat  = chosen.get("Item Category") or item_cat
                cond_type = chosen.get("Condition Type") or "ZPRM"
                cpi_flag  = chosen.get("CPI Eligible", "Yes")
            elif mat_code in master_lookup:
                dict_desc = picked_desc or master_lookup[mat_code].get("Description") or ""

        if dict_desc:
            desc_final = dict_desc
        else:
            desc_final = (desc_raw or "").strip()
            if len(desc_final) > 80:
                desc_final = desc_final[:77].rstrip() + "..."

        section_lower = (section or "").lower()
        desc_lower    = (desc_raw or "").lower()
        is_pass_through = any(k in (desc_lower + " " + section_lower)
                              for k in ("postage", "travel", "pass through",
                                        "pass-through", "third party",
                                        "third-party"))

        if is_pass_through and price_val is None:
            cond_type = ""

        if item_cat == "ZINL" or (is_pass_through and price_val is None):
            cpi_value = ""
        else:
            cpi_value = "CPI" if str(cpi_flag).lower().startswith("y") else ""

        if item_cat == "ZINL":
            row_start, row_end = start_date, last_day_start
            bill_plan = start_date
        else:
            row_start, row_end = start_date, end_date
            bill_plan = ""

        min_fee = _extract_minimum_fee(cond)
        cust_note = cust_note_base
        if min_fee is not None:
            cust_note = f"*PMIN {min_fee:.2f}*{cust_note_base}"

        if checkbox is True:
            checkbox_label = "Picked"
        elif checkbox is False:
            checkbox_label = "Not Picked"
        else:
            checkbox_label = "Does Not Exist"

        row = blank_row()
        row["MATERIAL"]        = mat_code or ""
        row["Desc"]            = desc_final
        row["Qty"]             = it.get("quantity") or 1
        row["Item Cat"]        = item_cat or ""
        row["Start Date"]      = row_start
        row["End Date"]        = row_end
        row["Cond.Type"]       = cond_type if mat_code else ""
        row["Price"]           = "" if price_val is None else price_val
        row["Currency"]        = "USDN" if price_val is not None else ""
        row["Cust. Mat. Num."] = cust_note
        row["Mat.Grp.2"]       = cpi_value if mat_code else ""
        row["Contract signed"] = signing_date
        row["Bill Plan Date"]  = bill_plan
        row["Checkbox"]        = checkbox_label
        row["Contract Page"]   = page_n if page_n is not None else ""
        row["Price Status"]    = price_status
        row["Source Contract"] = src_pdf

        rows.append(row)
        manual_codes = {"CUPR0754", "CUPRREPORTFLAT"}
        manual_flag = (min_fee is not None) or (mat_code in manual_codes)
        meta.append({
            "source_dict":      source_dict,
            "has_min":          (min_fee is not None),
            "manual_flag":      manual_flag,
            "no_match":         (not mat_code),
            "is_pass_through":  is_pass_through and price_val is None,
            "checkbox_not_picked": (checkbox is False),
        })

    return rows, meta, full_cols


# ─────────────────────────────────────────────────────────────────────────────
# Save with colour coding (verbatim from the script).
# ─────────────────────────────────────────────────────────────────────────────

FILL_GREEN     = PatternFill("solid", fgColor="C6EFCE")
FILL_BLUE      = PatternFill("solid", fgColor="BDD7EE")
FILL_ORANGE    = PatternFill("solid", fgColor="F4B084")
FILL_GREY      = PatternFill("solid", fgColor="D9D9D9")
FILL_PINK      = PatternFill("solid", fgColor="F8CBAD")
FILL_HEADER    = PatternFill("solid", fgColor="808080")
FILL_MAT_PMIN  = PatternFill("solid", fgColor="C65911")


def _pick_fill(m: dict) -> PatternFill:
    if m["no_match"]:            return FILL_ORANGE
    if m["checkbox_not_picked"]: return FILL_PINK
    if m["source_dict"] == "fallback": return FILL_BLUE
    if m["is_pass_through"]:     return FILL_GREY
    return FILL_GREEN


def _safe_write(df: pd.DataFrame, path: Path) -> Path:
    target = path
    idx = 1
    while True:
        try:
            df.to_excel(str(target), index=False)
            return target
        except PermissionError:
            target = path.with_name(f"{path.stem} ({idx}){path.suffix}")
            idx += 1
            if idx > 20:
                raise


def _colour_workbook(final_path: Path, full_cols: list[str],
                     row_meta: list[dict]) -> None:
    wb = load_workbook(str(final_path))
    ws = wb.active
    n_cols = len(full_cols)
    try:
        material_col = full_cols.index("MATERIAL") + 1
    except ValueError:
        material_col = None

    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).fill = FILL_HEADER
    for i, m in enumerate(row_meta):
        fill = _pick_fill(m)
        excel_row = i + 2
        for c in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=c).fill = fill
        if material_col is not None and m.get("manual_flag"):
            ws.cell(row=excel_row, column=material_col).fill = FILL_MAT_PMIN
    wb.save(str(final_path))
    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# PDF discovery — LATEST MASTER AGREEMENT ONLY (no year threshold).
# ─────────────────────────────────────────────────────────────────────────────
#
# MNR setup operates on the master agreement that's currently in force, not on
# its amendments, SOWs, or order forms (those reference the master's rates and
# are not the source of truth for SAP setup). And among multiple masters in a
# client folder (renewals, restated masters), only the latest one matters.
#
# This agent has NO year cutoff — even a master signed in 2008 is in scope if
# it's the latest one available. The chatbot's PDF-year cutoff (which gates
# Fee Description and CPI) is intentionally ignored here.
#
# Resolution chain, in priority order:
#   1. Hierarchy agent's per-client Excel — authoritative, uses Contract_Type.
#   2. Engagement Overview agent's per-client Excel — has 'Contract Type',
#      'Document Type', and 'Type of Agreement' columns, each of which can
#      classify a contract.
#   3. Filename heuristic — broad lowercase match for "master agreement" /
#      "master-agreement" / "master_agreement" / "msa" / "master".
# At each step the latest (by effective/signed date) wins.

# Substrings that *positively* identify a master agreement. Matched in
# lowercase against filenames, hierarchy Contract_Type, engagement-overview
# Contract Type / Document Type / Type of Agreement values.
_MASTER_TOKENS = (
    "master agreement", "master-agreement", "master_agreement",
    "masteragreement",                 # 'MasterAgreement' (Document Type in EO)
    "msa",
    "master",                          # broadest — catches 'Master Services
                                       #            Agreement', 'Master', etc.
)

# Substrings that NEGATE a master match — used to keep filename and engagement-
# overview fallbacks from picking up amendments, SOWs, addenda, order forms.
_NON_MASTER_TOKENS = (
    "amendment", "addendum", "sow", "statement of work",
    "statement-of-work", "order form", "order-form", "order_form",
    "schedule",
)


def _looks_like_master(text: str) -> bool:
    """Case-insensitive 'is this a master agreement?' classifier. Returns
    True iff `text` contains a master-token AND no non-master-token. Used
    by both the filename heuristic and the engagement-overview fallback."""
    if not text:
        return False
    low = str(text).lower()
    if any(tok in low for tok in _NON_MASTER_TOKENS):
        return False
    return any(tok in low for tok in _MASTER_TOKENS)


def _coerce_date(v) -> Optional[date]:
    """Best-effort coercion of any pandas/excel date-ish value to a python
    date. Returns None when nothing usable can be extracted."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, date):
        return v
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.notna(ts):
            return ts.date()
    except Exception:
        pass
    # Last resort — a "5-1-2011"-style string buried inside a longer value.
    s = str(v)
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        try:
            mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except Exception:
            return None
    return None


def _latest_master_from_hierarchy(client_name: str, folder: Path,
                                  allow: Optional[set],
                                  log: Callable[[str], None]
                                  ) -> Optional[tuple[str, date]]:
    """Hierarchy agent output: Output/<Client>/contracts_hierarchy.xlsx ->
    Sheet1. Filters Contract_Type for master-like values via
    `_looks_like_master`, picks the row with the latest Effective_Date
    (or Signed_Date / filename date as fallbacks)."""
    hxl = OUTPUT_DIR / client_name / "contracts_hierarchy.xlsx"
    if not hxl.exists():
        return None
    try:
        df = pd.read_excel(str(hxl), sheet_name="Sheet1")
    except Exception as e:
        log(f"  ⚠ Could not read {hxl.name}: {e}")
        return None
    if df.empty or "Contract_Type" not in df.columns or "Filename" not in df.columns:
        return None

    masters = df[df["Contract_Type"].astype(str).map(_looks_like_master)].copy()
    if masters.empty:
        return None

    candidates: list[tuple[str, date]] = []
    for _, row in masters.iterrows():
        fname = str(row.get("Filename") or "").strip()
        if not fname:
            continue
        if allow is not None and fname not in allow:
            continue
        # Skip restated masters that the hierarchy agent explicitly retired
        # in favour of a newer one.
        dup_of = row.get("Duplicate_Of")
        if isinstance(dup_of, str) and dup_of.strip():
            continue
        if not (folder / fname).exists():
            continue
        d = (_coerce_date(row.get("Effective_Date"))
             or _coerce_date(row.get("Signed_Date"))
             or _signing_date_from_filename(fname))
        if d is None:
            continue
        candidates.append((fname, d))

    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[1], t[0]), reverse=True)
    return candidates[0]


def _latest_master_from_engagement_overview(client_name: str, folder: Path,
                                            allow: Optional[set],
                                            log: Callable[[str], None]
                                            ) -> Optional[tuple[str, date]]:
    """Engagement Overview agent output: Output/<Client>/
    engagement_overview_output.xlsx. Tries 'Contract Type' first, then
    'Document Type', then 'Type of Agreement' — whichever first classifies
    the row as a master via `_looks_like_master`. Picks the latest by
    Effective Date / Contract Effective Date / filename date."""
    eox = OUTPUT_DIR / client_name / "engagement_overview_output.xlsx"
    if not eox.exists():
        return None
    try:
        df = pd.read_excel(str(eox))
    except Exception as e:
        log(f"  ⚠ Could not read {eox.name}: {e}")
        return None
    if df.empty or "Filename" not in df.columns:
        return None

    type_cols = [c for c in ("Contract Type", "Document Type",
                             "Type of Agreement") if c in df.columns]
    if not type_cols:
        return None

    def _row_is_master(row) -> bool:
        # ANY of the type columns classifying it as master is enough.
        return any(_looks_like_master(row.get(c)) for c in type_cols)

    masters = df[df.apply(_row_is_master, axis=1)].copy()
    if masters.empty:
        return None

    candidates: list[tuple[str, date]] = []
    for _, row in masters.iterrows():
        fname = str(row.get("Filename") or "").strip()
        if not fname:
            continue
        if allow is not None and fname not in allow:
            continue
        if not (folder / fname).exists():
            continue
        d = (_coerce_date(row.get("Effective Date"))
             or _coerce_date(row.get("Contract Effective Date"))
             or _signing_date_from_filename(fname))
        if d is None:
            continue
        candidates.append((fname, d))

    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[1], t[0]), reverse=True)
    return candidates[0]


def _latest_master_from_filename(folder: Path, allow: Optional[set]
                                 ) -> Optional[tuple[str, date]]:
    """Filename-only heuristic. No year filter — even an ancient master is
    in scope if nothing newer qualifies."""
    if not folder.exists():
        return None
    candidates: list[tuple[str, date]] = []
    for p in sorted(folder.glob("*.pdf")):
        if allow is not None and p.name not in allow:
            continue
        if not _looks_like_master(p.name):
            continue
        d = _signing_date_from_filename(p.name)
        if d is None:
            continue
        candidates.append((p.name, d))
    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[1], t[0]), reverse=True)
    return candidates[0]


def _find_latest_master(client_name: str, folder: Path,
                        contracts: Optional[list],
                        log: Callable[[str], None]
                        ) -> Optional[tuple[str, date]]:
    """Resolve the single (filename, signing_date) the MNR agent should
    process — the latest master agreement for this client. Tries three
    classifiers in order; the first one that returns a master wins.
    Returns None if no master can be identified at all."""
    allow = set(contracts) if contracts else None

    pick = _latest_master_from_hierarchy(client_name, folder, allow, log)
    if pick is not None:
        log(f"  Latest master (from Hierarchy): {pick[0]}  (effective {pick[1]})")
        return pick

    log("  Hierarchy did not classify a master — trying Engagement Overview.")
    pick = _latest_master_from_engagement_overview(client_name, folder, allow, log)
    if pick is not None:
        log(f"  Latest master (from Engagement Overview): {pick[0]}  "
            f"(effective {pick[1]})")
        return pick

    log("  Engagement Overview did not classify a master either — falling "
        "back to a filename heuristic.")
    pick = _latest_master_from_filename(folder, allow)
    if pick is not None:
        log(f"  Latest master (from filename): {pick[0]}  (signed {pick[1]})")
    return pick


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point — `run`
# ─────────────────────────────────────────────────────────────────────────────

def run(
    client_name: str,
    api_key: str = "",
    model: str = "",
    progress_callback: Optional[Callable[[str], None]] = None,
    contracts: Optional[list] = None,
    core: str = "",
    dictionary_path: Optional[Path] = None,
    force: bool = False,
    **_legacy_kwargs,
) -> dict:
    """MNR Template Agent — full 3-stage pipeline per client.

    Behaviour mirrors MNR_Setup_vRG_KeplerCannon_Apr2026.py. The only change
    is that Stage 1 / Stage 2 caches are namespaced per CLIENT and per PDF
    under  Output/<Client>/.mnr_cache/, so cache files can never collide
    across clients and no manual deletion is required between runs. Pass
    `force=True` to wipe the per-client cache and re-extract from scratch.

    Returns the standard agent result dict:
        {status, client, rows, output}  with status in
        {"complete", "no_pdfs", "no_catalog", "no_master", "no_template",
         "error"}.
    """
    log = progress_callback or (lambda m: None)
    api_key = api_key or MNR_API_KEY
    model_name = model or MNR_MODEL

    log(f"━━━ MNR Template {_VERSION} ━━━")
    log(f"Model: {model_name}")
    log("Scope: latest MASTER AGREEMENT only — amendments, SOWs, order "
        "forms, and older masters are intentionally excluded. No year "
        "threshold applied.")

    folder = client_input_dir(client_name, core)
    if not folder.exists():
        log(f"  Input folder does not exist: {folder}")
        return {"status": "no_pdfs", "client": client_name, "rows": 0,
                "output": ""}

    latest = _find_latest_master(client_name, folder, contracts, log=log)
    if latest is None:
        log("  No master agreement found for this client (Hierarchy, "
            "Engagement Overview, and filename heuristic all came up empty). "
            "Run the Hierarchy and/or Engagement Overview agents first so "
            "contracts are classified, then re-run.")
        return {"status": "no_master_agreement", "client": client_name,
                "rows": 0, "output": ""}

    pairs = [latest]   # exactly one PDF — keeps the rest of the loop intact

    # Locate the Frequently Used catalog (PRIMARY dictionary).
    freq_path = _find_freq_catalog(core)
    if freq_path is None:
        log(f"  ⚠ Could not find '{MNR_FREQ_CATALOG_NAME}' in Input/{core}/ "
            f"or {BASE_DIR}. The matcher needs this primary dictionary.")
        return {"status": "no_catalog", "client": client_name, "rows": 0,
                "output": ""}
    log(f"  Primary dictionary: {freq_path.name}")
    _ensure_cpi_column(freq_path, log)
    freq_df = pd.read_excel(str(freq_path)).fillna("").astype(str)

    # Locate the Portico master (FALLBACK dictionary). Uses the existing
    # Core dictionary resolution, or accepts an explicit override.
    master_path = Path(dictionary_path) if dictionary_path else default_dictionary_for(core)
    if not master_path or not Path(master_path).exists():
        log(f"  ⚠ Could not find a master dictionary for Core={core!r}.")
        return {"status": "no_master", "client": client_name, "rows": 0,
                "output": ""}
    log(f"  Fallback dictionary: {Path(master_path).name}")

    # Master sheet — same fallback chain as config.DICTIONARY_SHEET_CANDIDATES.
    from config import DICTIONARY_SHEET_CANDIDATES
    master_df = None
    for sheet in DICTIONARY_SHEET_CANDIDATES:
        try:
            master_df = pd.read_excel(str(master_path), sheet_name=sheet)
            log(f"    using sheet '{sheet}' from master dictionary.")
            break
        except Exception:
            continue
    if master_df is None:
        log(f"  ⚠ Could not read any of {DICTIONARY_SHEET_CANDIDATES} "
            f"from {master_path.name}.")
        return {"status": "no_master", "client": client_name, "rows": 0,
                "output": ""}
    master_df = master_df.fillna("").astype(str)
    if "Material Code" not in master_df.columns or "Description" not in master_df.columns:
        log(f"  ⚠ Master dictionary missing 'Material Code'/'Description' "
            f"columns; cannot match against fallback.")
        return {"status": "no_master", "client": client_name, "rows": 0,
                "output": ""}
    master_df = master_df[["Material Code", "Description"]].copy()

    primary_records = freq_df[[
        "Material Code", "Description", "Item Category",
        "Condition Type", "CPI Eligible"
    ]].to_dict(orient="records")
    fallback_records = master_df.to_dict(orient="records")
    master_lookup = {r["Material Code"]: r for r in fallback_records}

    # Resolve checkbox-reference image.
    checkbox_ref = Path(MNR_CHECKBOX_REF)
    if not checkbox_ref.exists():
        log(f"  ⚠ Checkbox reference image missing at {checkbox_ref}; "
            "Stage 1 still runs but checkbox detection accuracy will drop.")

    # Resolve MNR header template (optional).
    tmpl = _find_template(client_name, core)
    if tmpl is not None:
        try:
            template_df = pd.read_excel(str(tmpl))
            output_columns = list(template_df.columns)
            log(f"  Template: {tmpl.name} ({len(output_columns)} columns)")
        except Exception as e:
            log(f"  ⚠ Failed to read template {tmpl.name}: {e} — using "
                "default column header.")
            output_columns = list(_DEFAULT_TEMPLATE_COLUMNS)
    else:
        log(f"  No '{MNR_TEMPLATE_NAME}' template found — using default "
            f"column header ({len(_DEFAULT_TEMPLATE_COLUMNS)} cols).")
        output_columns = list(_DEFAULT_TEMPLATE_COLUMNS)

    if force:
        n = clear_caches(client_name)
        if n:
            log(f"  force=True: cleared {n} cached file(s) for this client.")

    ai_client = make_client(api_key)

    # Process each qualifying PDF and accumulate rows across the client.
    all_rows: list[dict] = []
    all_meta: list[dict] = []
    full_cols: list[str] = []
    n_extracted = 0
    n_matched = 0

    for fname, signing_date in pairs:
        pdf_path = str(folder / fname)
        pdf_stem = Path(fname).stem
        log(f"\n  ▸ {fname}  (signed {signing_date})")

        # Stage 1
        items = _stage1_extract(
            ai_client, pdf_path,
            cache_path=_stage1_cache(client_name, pdf_stem),
            checkbox_ref_path=checkbox_ref,
            model_name=model_name,
            log=log,
        )
        items = _drop_legal_clause_items(
            items, dropped_path=_dropped_path(client_name, pdf_stem), log=log
        )
        for it in items:
            it["_source_pdf"] = fname
        n_extracted += len(items)
        if not items:
            log("    No fee items survived Stage 1 + legal filter.")
            continue

        # Stage 2
        matched = _stage2_match(
            ai_client, items,
            primary_records=primary_records,
            fallback_records=fallback_records,
            cache_path=_stage2_cache(client_name, pdf_stem),
            model_name=model_name,
            log=log,
        )
        n_matched += sum(1 for m in matched if (m or {}).get("material_code"))

        # Stage 3
        rows, meta, cols = _build_mnr_rows(
            items, matched, signing_date, output_columns,
            primary_records=primary_records,
            master_lookup=master_lookup,
        )
        if not full_cols:
            full_cols = cols
        all_rows.extend(rows)
        all_meta.extend(meta)

    if not all_rows:
        log("\n  No MNR rows produced (every PDF yielded zero fee items).")
        return {"status": "complete", "client": client_name, "rows": 0,
                "output": ""}

    output_df = pd.DataFrame(all_rows, columns=full_cols)
    output_df = output_df.replace({"": pd.NA, "NaT": pd.NA})

    out_path = output_path(client_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    final_path = _safe_write(output_df, out_path)
    log(f"\n  Wrote {len(output_df)} row(s) → {final_path.name}")

    _colour_workbook(final_path, full_cols, all_meta)
    log("  Applied biller colour-coding (green/blue/orange/pink/grey, "
        "PMIN overlay).")

    return {
        "status": "complete",
        "client": client_name,
        "rows":   int(len(output_df)),
        "output": str(final_path),
        "stage1_items":  n_extracted,
        "stage2_matched": n_matched,
        "pdfs":          [p[0] for p in pairs],
    }


__all__ = ["run", "is_processed", "output_path", "clear_caches",
           "STAGE1_SYSTEM_PROMPT", "STAGE2_SYSTEM_PROMPT_TEMPLATE", "_VERSION"]
