"""
Material Validation Agent — post-processing layer over the Material Code
Matching Agent.

Pipeline:  Extraction → Matching → **Material Validation** → Final Output

WHAT IT DOES
------------
The Matching agent assigns each contract fee line a material code/description from
a *dictionary*. But billers use a many-to-many mapping: the same code means
different things for different clients, and the same description is billed under
different codes (~30k unique combinations — impossible to maintain in a dict). So
the dictionary match is, at best, a good guess at what was *actually billed*.

This agent re-anchors every row to what the client was REALLY billed, using that
client's own SAP invoice history in Snowflake
(``SAP_OTC_SD_INVOICE_CONSOLIDATED_VIEW``). It reads
``Output/<client>/material_match_output.xlsx`` and writes
``Output/<client>/validated_material_output.xlsx`` with the validated code, its
real invoice description, a GREEN/YELLOW/RED band, the four sub-scores, and a
human-readable reason per row.

THE ALGORITHM (the user's spec, implemented with the project's hard rules)
--------------------------------------------------------------------------
Per client, ONCE:
  0. SCOPE — resolve the client to its **Sold-To** number(s). One client = one
     sold-to (but may have several bill-tos). We match the client name against
     the sold-to NAME and also fold in the sold-tos of any matched bill-to.
     EVERY query is scoped to that sold-to so the output is always a code the
     client was actually billed. (Falls back to bill-to scope if the view has
     no sold-to column.)
  A. Per-code invoice stats over the scope (net/unit price spread, cadence,
     line counts, product hierarchy, sample descriptions).
  B. The scope's full (code, description) corpus — the universe of real billed
     descriptions used for lexical/semantic matching AND the probability signal.
  E. A GLOBAL exact-price scan (whole table) for the contract's exact amounts,
     each hit flagged in-scope / out-of-scope. Exact price is the strongest
     anchor for fixed-price items; out-of-scope-only hits become a "verify
     resolution" diagnostic — never a code change.

Per row:
  1. PRICE candidates — normalized price = net/qty (the contract price is often
     per-unit or tiered). Exact-price match first, then a ±25% buffer. Plus
     description/keyword candidates from the scope corpus, plus the incumbent.
     → a candidate pool (≤ POOL_CAP).
  2. SEMANTIC + CONTEXT — an LLM (see SEMANTIC_MODEL) ranks the pool using the
     fee item (55%) + matched description (35%) + section header (10%) against
     each candidate's REAL invoice descriptions, rewarding creative biller
     phrasings. A local TF-IDF score is the floor / fallback when the LLM is
     unavailable.  → narrow to top-5.
  3. PROBABILITY P(code | description, client) — from the scope corpus: given a
     description like this row's, how concentrated is the client's billing on
     each code.  → narrow to top-3.
  4. FREQUENCY — invoice cadence (monthly / one-time / annual, inferred from
     invoice-date spacing) vs the contract Frequency column (70%) and
     description hints like "OT" / "setup" / "monthly" (30%).
  5. FINAL = 0.30·price + 0.40·semantic_context + 0.15·probability
           + 0.15·frequency  → the chosen code (among the top-3).

HARD RULES (do not regress — see memory `validation-agent-design`)
------------------------------------------------------------------
  • Output ``new_material_code`` MUST be a code billed to THIS client (in scope).
    No invented codes. If no in-scope candidate exists, keep the incumbent and
    band RED.
  • Exact in-scope price match is king for fixed-price items — but a coincidental
    price match never beats a near-exact description match, and an exact-price
    anchor only counts at full weight when the code is also semantically
    plausible (or the amount is highly specific).
  • Credits are NEGATIVE net amounts → price matching is on |net| (a +32,133
    contract line must find the −32,133 credit). The LLM disambiguates a
    money-back credit from a same-priced service by wording + sign.
  • Tier pricing → ONE consistent code. Rows that are the same base fee split
    into volume tiers (the matcher appends the tier range as a trailing
    parenthetical) all get the single best-supported in-scope code.
  • NO sentence-transformers / HuggingFace, and this module must NOT import
    ``agents.extraction`` / ``agents.dna_extraction`` (they pull ST in
    transitively, which breaks on the firewalled Fiserv VDI). Semantics come
    from the LLM; lexical shortlisting/floor is local sklearn TF-IDF only.

DEGRADATION
-----------
When Snowflake is unreachable (off the VDI / no config / connector missing) the
agent SKIPS: it writes no file and returns a status the UI renders as a notice,
leaving the Matching output as the system of record. When only the LLM is
unreachable, it scores deterministically (TF-IDF + price + probability +
frequency) and notes that in ``validation_reason``.
"""

from __future__ import annotations

import json
import math
import os
import random
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Optional

import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import normalize
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import OUTPUT_DIR, VALIDATION_API_KEY, VALIDATION_MODEL
from fiserv_client import make_client
import snowflake_invoice as sf

# ── Final-score weights (the user's spec: price 30 · sem+ctx 40 · prob 15 · freq 15)
W_PRICE   = float(os.environ.get("VALIDATION_W_PRICE",   "0.30"))
W_SEMCTX  = float(os.environ.get("VALIDATION_W_SEMCTX",  "0.40"))
W_PROB    = float(os.environ.get("VALIDATION_W_PROB",    "0.15"))
W_FREQ    = float(os.environ.get("VALIDATION_W_FREQ",    "0.15"))

# Semantic query blend (fee item · matched description · section header).
W_ITEM_TXT  = 0.55
W_OLD_DESCP = 0.35
W_SECTION   = 0.10

# Frequency blend (contract Frequency column · description hint).
W_FREQ_COL  = 0.70
W_FREQ_DESC = 0.30

# Banding on the final composite.
GREEN_CUTOFF  = float(os.environ.get("VALIDATION_GREEN",  "0.80"))
YELLOW_CUTOFF = float(os.environ.get("VALIDATION_YELLOW", "0.60"))

# Price tolerance for the buffer band (±25%) — exact match is separate and exact.
NORM_PRICE_BUFFER = float(os.environ.get("VALIDATION_PRICE_BUFFER", "0.25"))

# Candidate-pool sizing (the user's "top 30 → top 5 → top 3" funnel).
POOL_CAP      = int(os.environ.get("VALIDATION_POOL_CAP", "30"))   # price+desc pool
LLM_CAND_CAP  = int(os.environ.get("VALIDATION_LLM_CANDS", "15"))  # sent to the LLM
TOP5          = int(os.environ.get("VALIDATION_TOP5", "5"))
TOP3          = int(os.environ.get("VALIDATION_TOP3", "3"))

# A near-exact match to the client's OWN invoice description anchors semantics
# (rule 6: description beats a coincidental price match).
DESC_ANCHOR_SIM = 0.85
# An exact-price code only gets the full price anchor when it's at least this
# semantically plausible (else a same-amount coincidence could hijack the row).
PRICE_ANCHOR_MIN_SEM = 0.40
# "Highly specific" = few distinct codes share this exact amount → the price is
# discriminative enough to anchor even with weak semantics.
SPECIFIC_PRICE_MAX_CODES = 2
# Semantic-confidence floor: when a pick has NO anchor (not exact-price, not a
# near-exact description) and weak semantics, the corroborating signals
# (probability / frequency) must NOT be allowed to make it look confident. Below
# this floor the reported score is scaled down so the band reflects the doubt.
SEM_CONF_FLOOR = float(os.environ.get("VALIDATION_SEM_FLOOR", "0.40"))

# ── Foundation-gateway LLM controls (memory: fiserv-vdi-deployment) ───────────
# Small batches + bounded concurrency + backoff because the gateway rate-limits
# (429) and returns EMPTY content when max_tokens is sent → default max_tokens=0.
SEMANTIC_MODEL   = os.environ.get("VALIDATION_SEMANTIC_MODEL") or VALIDATION_MODEL
LLM_WORKERS      = int(os.environ.get("VALIDATION_LLM_WORKERS", "3"))
LLM_BATCH        = int(os.environ.get("VALIDATION_LLM_BATCH", "3"))      # rows per call
LLM_RETRIES      = int(os.environ.get("VALIDATION_LLM_RETRIES", "6"))
LLM_RETRY_BASE   = float(os.environ.get("VALIDATION_LLM_RETRY_BASE", "3"))
LLM_RETRY_MAX    = float(os.environ.get("VALIDATION_LLM_RETRY_MAX", "60"))
LLM_MAXTOK       = int(os.environ.get("VALIDATION_LLM_MAXTOK", "0"))     # 0 = don't send

# How many distinct contract prices to scan globally in one pass (bounds Query E).
PRICE_SCAN_CAP   = int(os.environ.get("VALIDATION_PRICE_SCAN_CAP", "300"))

# Output columns appended to the matching sheet. The first block is the contract
# the frontend (chatbot.py table, context_builder system-of-record) depends on —
# do NOT rename those. The v_* / route / top_candidates are audit columns.
_NEW_COLUMNS = [
    "old_matched_description", "old_material_code", "old_match_confidence",
    "new_matched_description", "new_material_code", "new_validation_score",
    "confidence_band", "fallback_material_code", "validation_reason",
    "invoice_cadence", "validation_route",
    "v_price", "v_semantic_context", "v_probability", "v_frequency",
    "top_candidates",
]

_GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_RED_FILL    = PatternFill("solid", fgColor="FFC7CE")

# Legal-noise stripper (mirrors the extraction agent's clean()).
_LEGAL_NOISE = re.compile(
    r'\b(pursuant to|as defined in|hereinafter|section|schedule|exhibit'
    r'|agreement|contract|dated|effective|the foregoing|referred to as'
    r'|including but not limited to|subject to|in accordance with|attachment'
    r'|amendment|addendum)\b',
    flags=re.IGNORECASE,
)

# Tier parenthetical the matcher appends, e.g. "Card Fee (1 – 5,000 Active Users)".
_TIER_PAREN_RE = re.compile(r"\s*\([^)]*\)\s*$")
# A trailing volume range "… 5,001 – 7,500" with no parens.
_TIER_RANGE_RE = re.compile(r"[\-–—]\s*[\d,]+\s*[\-–—]\s*[\d,]+\s*$")


def _clean(text) -> str:
    return re.sub(r'\s+', ' ', _LEGAL_NOISE.sub(' ', str(text or ''))).strip()


def _clamp01(x) -> float:
    try:
        x = float(x)
    except (TypeError, ValueError):
        return 0.0
    if x != x:                       # NaN
        return 0.0
    return 0.0 if x < 0.0 else (1.0 if x > 1.0 else x)


def _f(v) -> Optional[float]:
    """Coerce Snowflake Decimal/None/str to float, or None (rejects NaN)."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if f == f else None


def _norm_code(c) -> str:
    """Case/space-insensitive code key for tolerant matching of LLM replies."""
    return re.sub(r"\s+", "", str(c or "").strip().upper())


def _cents(x) -> int:
    """Money as integer cents — an exact, float-repr-proof dict key for prices."""
    try:
        return int(round(abs(float(x)) * 100))
    except (TypeError, ValueError):
        return 0


# ── Public surface (mirrors the matcher; the frontend imports these) ──────────
def output_path(client_name: str) -> Path:
    return OUTPUT_DIR / client_name / "validated_material_output.xlsx"


def is_processed(client_name: str) -> bool:
    p = output_path(client_name)
    return p.exists() and p.stat().st_size > 1_000


# ── Local API retry (NO agents.dna_extraction import — that pulls in ST) ───────
def _call_with_retry(fn, log, retries: int = LLM_RETRIES,
                     base: float = LLM_RETRY_BASE, cap: float = LLM_RETRY_MAX):
    """Call ``fn`` with exponential backoff + jitter on rate-limit / transient
    errors. The Foundation gateway 429s under load; jitter de-syncs workers."""
    last = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:        # noqa: BLE001 — we re-raise non-transient
            last = e
            msg = str(e).lower()
            transient = ("429" in str(e) or "rate limit" in msg
                         or "too many requests" in msg or "timeout" in msg
                         or "timed out" in msg or "502" in msg or "503" in msg
                         or "504" in msg or "overloaded" in msg)
            if not transient or attempt == retries - 1:
                raise
            wait = min(cap, base * (2 ** attempt)) * (0.7 + 0.6 * random.random())
            log(f"  Rate/transient error (attempt {attempt + 1}/{retries}); "
                f"retrying in {wait:.1f}s…")
            time.sleep(wait)
    if last:
        raise last
    raise RuntimeError("API call failed after retries")


# ── Frequency canonicalisation + cadence inference ────────────────────────────
def _canon_freq(s) -> Optional[str]:
    """Map a contract Frequency / hint string to {One-Time, Monthly, Annual} or
    None. Handles the noisy real values ('Per Month', 'per member per month',
    'Per Transaction/Monthly', 'One Time', …)."""
    t = str(s or "").strip().lower()
    if not t or t == "nan":
        return None
    # Recurring beats one-time when both appear (e.g. "setup then monthly").
    if any(k in t for k in ("month", "/mo", "p/m", "per stmt", "recurring",
                            "per member", "per user", "per transaction",
                            "per item", "per account")):
        if "month" in t or "/mo" in t or "recurring" in t:
            return "Monthly"
    if any(k in t for k in ("annual", "annually", "yearly", "per year",
                            "/year", "p.a", " pa ", "per annum")):
        return "Annual"
    if "month" in t:
        return "Monthly"
    if any(k in t for k in ("one time", "one-time", "onetime", "set up",
                            "set-up", "setup", "implementation", "implement",
                            "upfront", "initial", "one off", "one-off",
                            " ot ", "install")):
        return "One-Time"
    return None


def _infer_cadence(st: Optional[dict]) -> str:
    """Infer a code's billing cadence from the spacing of its invoice dates."""
    if not st:
        return "Unknown"
    fd, ld = st.get("first_date"), st.get("last_date")
    n_inv = int(st.get("n_invoices") or 0)
    n_ym = int(st.get("n_year_months") or 0)
    if fd is None or ld is None or n_inv <= 1:
        return "One-Time"
    try:
        fd = pd.to_datetime(fd)
        ld = pd.to_datetime(ld)
        span_days = max((ld - fd).days, 0)
    except Exception:
        return "Unknown"
    if span_days < 45:
        return "One-Time"
    span_months = max(span_days / 30.44, 1.0)
    ratio = n_ym / span_months if span_months else 0.0
    avg_gap = span_days / max(n_inv - 1, 1)
    if ratio >= 0.6 or avg_gap <= 45:
        return "Monthly"
    if avg_gap >= 300:
        return "Annual"
    return "Unknown"


def _cadence_agree(a: Optional[str], b: str) -> float:
    if a is None or b in ("Unknown", ""):
        return 0.5                       # neutral when either side unknown
    if a == b:
        return 1.0
    ladder = {"One-Time": 0, "Monthly": 1, "Annual": 2}
    if a in ladder and b in ladder:
        return {1: 0.6, 2: 0.3}.get(abs(ladder[a] - ladder[b]), 0.0)
    return 0.0


def _freq_signal(contract_freq, item_text, cadence: str) -> float:
    """Frequency score: contract Frequency column (70%) + description hint (30%)
    vs the code's inferred cadence. Falls back to whichever side is present;
    neutral (0.5) when both are missing."""
    cf = _canon_freq(contract_freq)
    dh = _canon_freq(item_text)
    if cf is not None and dh is not None:
        return W_FREQ_COL * _cadence_agree(cf, cadence) + W_FREQ_DESC * _cadence_agree(dh, cadence)
    if cf is not None:
        return _cadence_agree(cf, cadence)
    if dh is not None:
        return _cadence_agree(dh, cadence)
    return 0.5


# ── Zero-priced + per-unit detection (ports of the proven helpers) ────────────
_ZERO_PRICE_KWS = (
    "included", "incl.", "incl ", "waiv", "prev paid", "prevpaid",
    "previously paid", "pre-paid", "prepaid", "no charge", "no-charge",
    "n/c", "n / c", "bundled", "complimentary", "at no cost", "free of charge",
)
_PER_UNIT_RE = re.compile(
    r"\bper\s+(user|account|acct|item|transaction|txn|trans|check|cheque|card|"
    r"statement|stmt|member|customer|location|branch|device|seat|licen[cs]e|call|"
    r"minute|click|unit|record|document|page|enrollment|loan|certificate|"
    r"appliance|mailbox|box|hour|gigabyte|gb|api|request|lookup|inquiry|alert|"
    r"message|sms|email|notice|deposit|withdrawal|payment|file|report|"
    r"envelope|kilobyte|subscription|application|end user)s?\b", re.I)


def _is_zero_priced(row, cleaned: Optional[float]) -> bool:
    raw  = str(row.get("Price", "") or "")
    item = str(row.get("Item", "") or "")
    blob = f"{raw} {item}".lower()
    if any(kw in blob for kw in _ZERO_PRICE_KWS):
        return True
    raw_s = raw.strip().lower().replace("$", "").replace(",", "").replace(" ", "")
    if raw_s in ("0", "0.0", "0.00", "0.000", "-"):
        return True
    if (cleaned is not None and abs(cleaned) < 0.005
            and raw.strip().lower() not in ("", "nan", "none")):
        return True
    return False


def _is_per_unit(freq, item) -> bool:
    return bool(_PER_UNIT_RE.search(f"{freq or ''} {item or ''}"))


# ── Tier grouping (rule 5 / user rule 1: tiers → one code) ────────────────────
def _base_fee_key(item, section) -> tuple:
    """Collapse an item to its base fee (drop the trailing tier parenthetical /
    range) + section, so volume tiers of one fee group together."""
    base = str(item or "").strip()
    base = _TIER_PAREN_RE.sub("", base).strip()
    base = _TIER_RANGE_RE.sub("", base).strip()
    return (re.sub(r"\s+", " ", base.lower()), str(section or "").strip().lower())


# ─────────────────────────────────────────────────────────────────────────────
# SCOPE — client → Sold-To number(s), the spine of "invoice-present" output.
# ─────────────────────────────────────────────────────────────────────────────
def _actual(col: str, info: dict) -> Optional[str]:
    if col in info.get("present", []):
        return info.get("aliased", {}).get(col, col)
    return None


def _resolve_scope(client_name: str, cfg: dict, info: dict, log) -> Optional[dict]:
    """Resolve the client to a Snowflake filter (SOLD-TO preferred, sold-to-name
    match, bill-to fallback). Returns ``{col, values, kind, label}`` or None.

    Delegates to ``snowflake_invoice.resolve_scope`` so this agent and the chat
    share ONE implementation — that's what guarantees they can't disagree on
    whether a code was billed to a client (the bug this consolidation fixes)."""
    return sf.resolve_scope(client_name, cfg, info, log=log)


def _scope_clause(scope: dict) -> tuple:
    """Return (sql_predicate, params) for a WHERE/CASE on the scope."""
    ph = ", ".join(["%s"] * len(scope["values"]))
    return f"{scope['col']} IN ({ph})", tuple(str(v) for v in scope["values"])


# ─────────────────────────────────────────────────────────────────────────────
# SNOWFLAKE AGGREGATION — three scoped queries + one global price scan / client.
# ─────────────────────────────────────────────────────────────────────────────
def _split_listagg(v) -> set:
    if not v:
        return set()
    return {p.strip() for p in str(v).split("||") if p and p.strip()}


def _fetch_scope_aggregates(scope: dict, cfg: dict, info: dict, log) -> dict:
    """Query A (per-code stats) + Query B (code/desc corpus), scoped. Returns the
    client's full billed-code universe — the candidate source and stats store."""
    tbl = sf._qualified_table(cfg)
    mat   = _actual("OTC_SIL_MATERIAL", info)
    mtext = _actual("OTC_SIL_MATERIAL_TEXT", info)
    net   = _actual("OTC_SIL_NET_AMOUNT", info)
    qty   = _actual("OTC_SIL_ORDER_QUANTITY", info)
    date  = _actual("OTC_SIH_INVOICE_DATE", info)
    doc   = _actual("OTC_SIH_INVOICE_DOCUMENT", info)
    prodh = _actual("OTC_SIL_PRODUCT_HIER_NAME", info)
    pctr  = _actual("OTC_SIL_PROFIT_CENTER_NAME", info)

    empty = {"by_code": {}, "total_lines": 0, "corpus": [], "code_descs": {}}
    if not mat or not mtext:
        log("  Invoice view lacks material/description columns — cannot validate.")
        return empty

    where, params = _scope_clause(scope)
    net_expr  = (f"TRY_TO_DECIMAL(REGEXP_REPLACE({net}::STRING,'[^0-9.-]',''),38,4)"
                 if net else None)
    qty_expr  = (f"TRY_TO_DECIMAL(REGEXP_REPLACE({qty}::STRING,'[^0-9.-]',''),38,4)"
                 if qty else None)
    unit_expr = (f"({net_expr})/NULLIF({qty_expr},0)" if (net_expr and qty_expr) else None)
    date_expr = f"TRY_TO_DATE({date}::STRING)" if date else None
    conn = sf._get_connection()

    # ── Query A — per material code ──────────────────────────────────────────
    sel = [("CODE", mat), ("N_LINES", "COUNT(*)"),
           ("N_DISTINCT_DESC", f"COUNT(DISTINCT {mtext})"),
           ("SAMPLE_DESC", f"MAX({mtext})")]
    if net_expr:
        sel += [("MIN_NET", f"MIN({net_expr})"), ("MAX_NET", f"MAX({net_expr})"),
                ("MEDIAN_NET", f"MEDIAN({net_expr})"),
                ("MIN_ABS_NET", f"MIN(ABS({net_expr}))"),
                ("MAX_ABS_NET", f"MAX(ABS({net_expr}))"),
                ("MEDIAN_ABS_NET", f"MEDIAN(ABS({net_expr}))")]
    if unit_expr:
        sel += [("MEDIAN_UNIT", f"MEDIAN(ABS({unit_expr}))"),
                ("MIN_UNIT", f"MIN(ABS({unit_expr}))"),
                ("MAX_UNIT", f"MAX(ABS({unit_expr}))")]
    if date_expr:
        sel += [("FIRST_DATE", f"MIN({date_expr})"), ("LAST_DATE", f"MAX({date_expr})"),
                ("N_YEAR_MONTHS", f"COUNT(DISTINCT TO_CHAR({date_expr},'YYYY-MM'))")]
    if doc:
        sel += [("N_INVOICES", f"COUNT(DISTINCT {doc})")]
    if prodh:
        sel += [("PROD_HIERS", f"LISTAGG(DISTINCT {prodh}, '||')")]
    if pctr:
        sel += [("PROFIT_CENTERS", f"LISTAGG(DISTINCT {pctr}, '||')")]
    cols_sql = ", ".join(f"{expr} AS {alias}" for alias, expr in sel)
    sql_a = (f"SELECT {cols_sql} FROM {tbl} "
             f"WHERE {where} AND {mat} IS NOT NULL GROUP BY {mat}")
    by_code: dict = {}
    cur = conn.cursor()
    cur.execute(sql_a, params)
    a_cols = [str(d[0]).upper() for d in (cur.description or [])]
    for row in cur.fetchall():
        r = dict(zip(a_cols, row))
        code = str(r.get("CODE") or "").strip()
        if not code:
            continue
        by_code[code] = {
            "n_lines": int(r.get("N_LINES") or 0),
            "n_distinct_desc": int(r.get("N_DISTINCT_DESC") or 0),
            "sample_desc": str(r.get("SAMPLE_DESC") or ""),
            "min_net": _f(r.get("MIN_NET")), "max_net": _f(r.get("MAX_NET")),
            "median_net": _f(r.get("MEDIAN_NET")),
            "min_abs_net": _f(r.get("MIN_ABS_NET")), "max_abs_net": _f(r.get("MAX_ABS_NET")),
            "median_abs_net": _f(r.get("MEDIAN_ABS_NET")),
            "median_unit": _f(r.get("MEDIAN_UNIT")),
            "min_unit": _f(r.get("MIN_UNIT")), "max_unit": _f(r.get("MAX_UNIT")),
            "first_date": r.get("FIRST_DATE"), "last_date": r.get("LAST_DATE"),
            "n_year_months": int(r.get("N_YEAR_MONTHS") or 0),
            "n_invoices": int(r.get("N_INVOICES") or 0),
            "prod_hiers": _split_listagg(r.get("PROD_HIERS")),
            "profit_centers": _split_listagg(r.get("PROFIT_CENTERS")),
        }
    total_lines = sum(v["n_lines"] for v in by_code.values())
    if total_lines == 0:
        return empty

    # ── Query B — distinct (code, description) corpus ────────────────────────
    sql_b = (f"SELECT {mat} AS CODE, {mtext} AS DESCR, COUNT(*) AS N_LINES FROM {tbl} "
             f"WHERE {where} AND {mat} IS NOT NULL AND {mtext} IS NOT NULL "
             f"AND LENGTH(TRIM({mtext})) > 0 GROUP BY {mat}, {mtext}")
    corpus: list = []
    code_descs: dict = {}
    cur = conn.cursor()
    cur.execute(sql_b, params)
    b_cols = [str(d[0]).upper() for d in (cur.description or [])]
    for row in cur.fetchall():
        r = dict(zip(b_cols, row))
        code = str(r.get("CODE") or "").strip()
        desc = str(r.get("DESCR") or "").strip()
        if not code or not desc:
            continue
        n = int(r.get("N_LINES") or 0)
        corpus.append({"code": code, "desc": desc, "n_lines": n})
        code_descs.setdefault(code, []).append((desc, n))
    for c in code_descs:                 # keep the 3 most-billed phrasings / code
        code_descs[c] = [d for d, _ in sorted(code_descs[c], key=lambda x: -x[1])[:3]]

    log(f"  Invoice history (scope): {total_lines} lines · {len(by_code)} codes · "
        f"{len(corpus)} distinct descriptions.")
    return {"by_code": by_code, "total_lines": total_lines, "corpus": corpus,
            "code_descs": code_descs}


def _fetch_exact_price_hits(prices: list, scope: dict, cfg: dict, info: dict,
                            log) -> dict:
    """GLOBAL exact-price scan (whole table) for the contract's |amounts|.

    Returns {rounded_abs_price: {"in": [recs], "out": [recs], "n_codes": int}}
    where each rec = {code, desc, signed_net, is_credit, n_lines}. ``in`` =
    billed within the client's scope (feeds candidates/anchor); ``out`` =
    billed only elsewhere (diagnostic — never changes a code).

    Matching is on ABS(net) so a +X contract line finds a −X credit. Bounded to
    PRICE_SCAN_CAP distinct amounts in ONE query."""
    out: dict = {}
    mat   = _actual("OTC_SIL_MATERIAL", info)
    mtext = _actual("OTC_SIL_MATERIAL_TEXT", info)
    net   = _actual("OTC_SIL_NET_AMOUNT", info)
    if not (mat and net):
        return out
    amounts = sorted({round(abs(p), 2) for p in prices if p and abs(p) >= 0.01})
    if not amounts:
        return out
    if len(amounts) > PRICE_SCAN_CAP:
        log(f"  ⚠ {len(amounts)} distinct prices > cap {PRICE_SCAN_CAP}; "
            f"scanning the {PRICE_SCAN_CAP} largest.")
        amounts = sorted(amounts, reverse=True)[:PRICE_SCAN_CAP]

    tbl = sf._qualified_table(cfg)
    net_expr = f"TRY_TO_DECIMAL(REGEXP_REPLACE({net}::STRING,'[^0-9.-]',''),38,4)"
    abs_expr = f"ROUND(ABS({net_expr}),2)"
    scope_pred, scope_params = _scope_clause(scope)
    in_list = ", ".join(f"{a:.2f}" for a in amounts)   # numeric literals, safe
    desc_sel = f"MAX({mtext})" if mtext else "''"
    sql = (
        f"SELECT {mat} AS CODE, {abs_expr} AS BUCKET, "
        f"CASE WHEN {scope_pred} THEN 1 ELSE 0 END AS IN_SCOPE, "
        f"COUNT(*) AS N_LINES, {desc_sel} AS SAMPLE_DESC, "
        f"MIN({net_expr}) AS MIN_NET, MAX({net_expr}) AS MAX_NET "
        f"FROM {tbl} WHERE {mat} IS NOT NULL AND {abs_expr} IN ({in_list}) "
        f"GROUP BY 1, 2, 3 "                       # ordinals: CODE, BUCKET, IN_SCOPE
        f"ORDER BY 3 DESC, 4 DESC LIMIT 20000")
    conn = sf._get_connection()
    cur = conn.cursor()
    try:
        cur.execute(sql, scope_params)        # params feed the CASE predicate
    except Exception as e:
        log(f"  ⚠ exact-price scan failed: {e}")
        return out
    cols = [str(d[0]).upper() for d in (cur.description or [])]
    code_seen: dict = {}
    for row in cur.fetchall():
        r = dict(zip(cols, row))
        code = str(r.get("CODE") or "").strip()
        bucket = _f(r.get("BUCKET"))
        if not code or bucket is None:
            continue
        key = _cents(bucket)
        in_scope = int(r.get("IN_SCOPE") or 0) == 1
        mn, mx = _f(r.get("MIN_NET")), _f(r.get("MAX_NET"))
        signed = mn if (mn is not None and mn < 0) else (mx if mx is not None else bucket)
        rec = {"code": code, "desc": str(r.get("SAMPLE_DESC") or ""),
               "signed_net": signed, "is_credit": (signed is not None and signed < 0),
               "n_lines": int(r.get("N_LINES") or 0)}
        b = out.setdefault(key, {"in": [], "out": [], "codes": set()})
        b["in" if in_scope else "out"].append(rec)
        b["codes"].add(code)
    for bucket, b in out.items():
        b["n_codes"] = len(b["codes"])
    n_in = sum(len(b["in"]) for b in out.values())
    n_out = sum(len(b["out"]) for b in out.values())
    log(f"  Exact-price scan: {len(out)} amount(s) matched — "
        f"{n_in} in-scope hit(s), {n_out} out-of-scope (diagnostic).")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# PRICE SIGNAL
# ─────────────────────────────────────────────────────────────────────────────
def _price_band(price: float, med, lo, hi) -> float:
    """Decay a price against a [lo, hi] band with median, plus the ±25% buffer."""
    if med is not None and abs(price - med) <= 0.02 * max(abs(med), 1.0):
        return 1.0
    if lo is not None and hi is not None and lo <= price <= hi:
        return 0.9
    buf = NORM_PRICE_BUFFER
    if lo is not None and hi is not None and lo * (1 - buf) <= price <= hi * (1 + buf):
        return 0.7
    if med is not None and abs(price - med) <= buf * max(abs(med), 1.0):
        return 0.7
    if lo is None or hi is None:
        return 0.5
    band = max(hi - lo, abs(med or 0.0), 1.0)
    dist = (lo - price) if price < lo else (price - hi)
    return max(0.0, 0.7 * (1.0 - dist / (3.0 * band)))


def _price_signal(price, st, is_zero: bool, per_unit: bool, is_exact: bool) -> float:
    """Per-candidate price score ∈ [0,1]. Exact in-scope match → 1.0 (king)."""
    if is_exact:
        return 1.0
    if is_zero:
        if not st:
            return 0.5
        if st.get("min_abs_net") is not None and abs(st["min_abs_net"]) < 0.5:
            return 0.85               # code has historically-free lines
        return 0.5                    # billed code, "included" is neutral
    if not st:
        return 0.5
    p = abs(price) if price else 0.0
    if p <= 0:
        return 0.5                    # price unknown → neutral
    if per_unit:
        med, lo, hi = st.get("median_unit"), st.get("min_unit"), st.get("max_unit")
        if med is None and lo is None and hi is None:
            return 0.5
    else:
        med, lo, hi = st.get("median_abs_net"), st.get("min_abs_net"), st.get("max_abs_net")
    return _price_band(p, med, lo, hi)


# ─────────────────────────────────────────────────────────────────────────────
# LLM SEMANTIC + CONTEXT RE-RANKER (Foundation API; truncation-tolerant)
# ─────────────────────────────────────────────────────────────────────────────
_RERANK_SYSTEM = (
    "You are a SAP material-code validation expert. Billers type invoice "
    "descriptions in their own words — often creative paraphrases of what a "
    "contract called the fee. For each contract fee line you get: the fee ITEM "
    "text, the prior MATCHED description, the contract SECTION header (a product/"
    "context hint — sometimes just legal boilerplate, ignore it then), the "
    "contract price & frequency, and a list of CANDIDATE material codes — each "
    "with REAL invoice descriptions billed to THIS client, a sample net amount "
    "(negative = a credit / money returned), and how many lines used it.\n\n"
    "Weigh meaning ~80% and product/section context ~20%. Reward plausible "
    "creative phrasings. A NEGATIVE-net 'credit/rebate' line (money back) is NOT "
    "the same as a positive-net service sold at the same amount — disambiguate by "
    "wording and sign, not amount alone. You may ONLY choose from the candidate "
    "codes given — never invent a code.\n\n"
    "For each row, rank the candidates you find plausible (best first) with a "
    "relevance 0-100. Return STRICT JSON, no markdown:\n"
    '{ "results": [ { "row_id": <int>, "ranked": [ {"code": "<cand>", '
    '"relevance": <0-100>} ], "reason": "<one concise sentence>" } ] }'
)


def _parse_llm_results(raw: str) -> list:
    """Salvage the results array from a possibly fenced / truncated reply.

    The Foundation gateway truncates long JSON, which silently dropped every
    pick in the past. Strategy: strip fences → json.loads → first{…}last} →
    regex-scrape complete ``{...}`` objects that contain a row_id."""
    if not raw:
        return []
    s = raw.strip()
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z]*\n?", "", s)
        s = re.sub(r"```$", "", s).strip()
    # 1) clean parse
    try:
        obj = json.loads(s)
        if isinstance(obj, dict) and isinstance(obj.get("results"), list):
            return obj["results"]
        if isinstance(obj, list):
            return obj
    except json.JSONDecodeError:
        pass
    # 2) outermost braces
    a, b = s.find("{"), s.rfind("}")
    if a >= 0 and b > a:
        try:
            obj = json.loads(s[a:b + 1])
            if isinstance(obj, dict) and isinstance(obj.get("results"), list):
                return obj["results"]
        except json.JSONDecodeError:
            pass
    # 3) scrape per-row objects from a truncated stream
    out = []
    for m in re.finditer(r'\{[^{}]*"row_id"[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', s, re.S):
        frag = m.group(0)
        try:
            out.append(json.loads(frag))
        except json.JSONDecodeError:
            rid = re.search(r'"row_id"\s*:\s*(\d+)', frag)
            codes = re.findall(r'"code"\s*:\s*"([^"]+)"\s*,\s*"relevance"\s*:\s*(\d+)', frag)
            if rid:
                out.append({"row_id": int(rid.group(1)),
                            "ranked": [{"code": c, "relevance": int(rl)} for c, rl in codes]})
    return out


def _llm_semantic_rank(client, payloads: list, log) -> dict:
    """payloads: per-row dicts. Returns {row_id: {ranked:[(code, rel0to1)], reason}}.
    Enhancement only — any failure leaves the row to deterministic scoring."""
    out: dict = {}
    if not payloads:
        return out
    batches = [payloads[i:i + LLM_BATCH] for i in range(0, len(payloads), LLM_BATCH)]

    def _one(batch):
        user = json.dumps({"rows": batch}, default=str)
        base_kwargs = dict(
            model=SEMANTIC_MODEL,
            messages=[{"role": "system", "content": _RERANK_SYSTEM},
                      {"role": "user", "content": user}],
            temperature=0,
        )
        if LLM_MAXTOK > 0:
            base_kwargs["max_tokens"] = LLM_MAXTOK

        def _send(kw):
            return _call_with_retry(
                lambda: client.chat.completions.create(**kw), log)

        try:
            resp = _send(base_kwargs)
        except Exception as e:
            log(f"  ⚠ rerank batch failed: {e}")
            return []
        raw, finish = _content_of(resp)
        if not raw:
            # Empty content correlates with sending temperature/max_tokens to the
            # gateway — retry once bare, and log the finish_reason for diagnosis.
            log(f"  ⚠ empty rerank content (finish={finish}); retrying bare.")
            bare = dict(model=SEMANTIC_MODEL, messages=base_kwargs["messages"])
            try:
                resp = _send(bare)
                raw, finish = _content_of(resp)
            except Exception as e:
                log(f"  ⚠ bare retry failed: {e}")
                return []
        return _parse_llm_results(raw)

    with ThreadPoolExecutor(max_workers=LLM_WORKERS) as ex:
        futs = [ex.submit(_one, b) for b in batches]
        for fut in as_completed(futs):
            for r in (fut.result() or []):
                try:
                    rid = int(r.get("row_id"))
                except (TypeError, ValueError):
                    continue
                ranked = []
                for c in (r.get("ranked") or []):
                    if not isinstance(c, dict):
                        continue
                    code = str(c.get("code") or "").strip()
                    if not code:
                        continue
                    relv = _f(c.get("relevance"))
                    rel = _clamp01((relv / 100.0) if relv is not None else 0.0)
                    ranked.append((code, rel))
                out[rid] = {"ranked": ranked, "reason": str(r.get("reason") or "").strip()}
    applied = sum(1 for v in out.values() if v["ranked"])
    log(f"  LLM ranks applied: {applied}/{len(payloads)}")
    return out


def _content_of(resp) -> tuple:
    raw, finish = "", None
    if getattr(resp, "choices", None):
        ch = resp.choices[0]
        finish = getattr(ch, "finish_reason", None)
        msg = getattr(ch, "message", None)
        raw = (getattr(msg, "content", None) or "").strip() if msg else ""
    return raw, finish


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY
# ─────────────────────────────────────────────────────────────────────────────
def run(client_name: str, api_key: str = "", progress_callback=None,
        contracts: Optional[list] = None, core: str = "") -> dict:
    _raw_log = progress_callback or (lambda m: None)
    _lock = threading.Lock()

    def log(msg: str) -> None:
        # Thread-safe; the rerank runs in worker threads where a Streamlit
        # callback raises NoSessionContext — swallow and fall back to stdout.
        try:
            with _lock:
                _raw_log(msg)
        except Exception:
            try:
                print(msg)
            except Exception:
                pass

    def _skip(reason: str, detail: str = "") -> dict:
        log(f"⏭ Skipped — {reason}")
        return {"status": "no_snowflake" if reason.startswith("Snowflake")
                else _status_for(reason),
                "client": client_name, "rows": 0, "output": "",
                "reason": reason, "detail": detail}

    # 1) Snowflake gate — skip (no file) when invoice data is unreachable.
    if not sf.snowflake_available():
        return _skip("Snowflake connector not installed",
                     "pip install snowflake-connector-python")
    try:
        cfg = sf._load_cfg()
    except Exception as e:
        return _skip("Snowflake config not found", str(e))
    try:
        info = sf._discover_columns(cfg)
    except Exception as e:
        return _skip("Snowflake connection failed", str(e))
    if info.get("missing_required"):
        return _skip("Snowflake view missing required columns",
                     ", ".join(info["missing_required"]))

    # 2) Read the Matching Agent output.
    in_path = OUTPUT_DIR / client_name / "material_match_output.xlsx"
    if not in_path.exists():
        return {"status": "no_matching", "client": client_name, "rows": 0,
                "output": "", "reason": "material_match_output.xlsx not found"}
    try:
        df = pd.read_excel(str(in_path))
    except Exception as e:
        return {"status": "no_matching", "client": client_name, "rows": 0,
                "output": "", "reason": f"could not read matching output: {e}"}
    if df.empty:
        return {"status": "no_items", "client": client_name, "rows": 0,
                "output": "", "reason": "matching output is empty"}

    section_field = ("Normalized_Section" if "Normalized_Section" in df.columns
                     else "Section_Header")
    price_field = "Cleaned Price" if "Cleaned Price" in df.columns else "Price"
    md_col   = "Matched Description" if "Matched Description" in df.columns else None
    code_col = "Material Code" if "Material Code" in df.columns else None
    conf_col = "Match Confidence" if "Match Confidence" in df.columns else None
    log(f"Loaded {len(df)} matched row(s).")

    # 3) SCOPE — client → sold-to (or bill-to fallback). Every query is scoped.
    log(f"Resolving scope for '{client_name}'…")
    try:
        scope = _resolve_scope(client_name, cfg, info, log)
    except Exception as e:
        return _skip("Snowflake connection failed", f"scope: {e}")
    if not scope:
        return {"status": "no_history", "client": client_name, "rows": 0,
                "output": "", "reason": f"no SAP sold-to/bill-to matched '{client_name}'"}

    # 4) One-time scoped aggregation (Query A + B) + global exact-price scan (E).
    log("Aggregating invoice history from Snowflake…")
    try:
        agg = _fetch_scope_aggregates(scope, cfg, info, log)
    except Exception as e:
        return _skip("Snowflake query failed", str(e))
    if agg["total_lines"] == 0:
        return {"status": "no_history", "client": client_name, "rows": 0,
                "output": "", "reason": "no invoice lines for the resolved scope"}

    contract_prices = []
    for v in df[price_field].astype(object):
        p = sf._safe_float(v)
        if p:
            contract_prices.append(p)
    try:
        price_hits = _fetch_exact_price_hits(contract_prices, scope, cfg, info, log)
    except Exception as e:
        log(f"  ⚠ exact-price scan error: {e}")
        price_hits = {}

    # 5) Build the local lexical index over the scope corpus (TF-IDF char-n-gram;
    #    NO sentence-transformers). Used for candidate generation, the semantic
    #    floor, and the probability signal.
    corpus = agg["corpus"]
    corpus_codes = [c["code"] for c in corpus]
    corpus_nlines = [max(int(c["n_lines"]), 1) for c in corpus]
    corpus_text = [_clean(f"{c['desc']}") for c in corpus]
    tfidf = corpus_M = None
    if corpus_text:
        tfidf = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4),
                                min_df=1, sublinear_tf=True)
        corpus_M = normalize(tfidf.fit_transform(corpus_text))
    by_code = agg["by_code"]
    total_lines = agg["total_lines"]

    def _query_vec(item_clean, md_clean, section_clean):
        if tfidf is None:
            return None
        parts = [(W_ITEM_TXT, item_clean), (W_OLD_DESCP, md_clean),
                 (W_SECTION, section_clean)]
        qv = None
        for w, txt in parts:
            if not txt:
                continue
            v = normalize(tfidf.transform([txt])) * w
            qv = v if qv is None else qv + v
        return normalize(qv) if qv is not None else None

    def _lex_sims(qv):
        """Per-corpus-row cosine sims for a query vector (or None)."""
        if qv is None or corpus_M is None:
            return None
        return (qv @ corpus_M.T).toarray().ravel()

    _sim_cache: dict = {}

    def _best_sim_by_code(sims):
        """Collapse per-row sims to the best sim per code."""
        best: dict = {}
        if sims is None:
            return best
        for i, code in enumerate(corpus_codes):
            s = float(sims[i])
            if s > best.get(code, -1.0):
                best[code] = s
        return best

    def _probability(sims, cand_codes):
        """P(code | description, client) over the candidate set: line-mass of
        descriptions like this row's, concentrated by code. Falls back to
        P(code|client) when no description matches."""
        cand = set(cand_codes)
        mass: dict = {c: 0.0 for c in cand}
        if sims is not None:
            for i, code in enumerate(corpus_codes):
                if code in cand:
                    s = float(sims[i])
                    if s >= 0.30:
                        mass[code] += s * corpus_nlines[i]
        tot = sum(mass.values())
        if tot > 0:
            return {c: mass[c] / tot for c in cand}
        # Fallback: prevalence within the client (line share among candidates).
        nl = {c: (by_code.get(c, {}).get("n_lines", 0)) for c in cand}
        ntot = sum(nl.values())
        if ntot > 0:
            return {c: nl[c] / ntot for c in cand}
        return {c: 0.0 for c in cand}

    # 6) Per-row candidate generation + deterministic pre-scoring.
    rows_meta: dict = {}      # idx -> dict(item_clean, price, ...)
    gpt_payloads: list = []

    for idx, row in df.iterrows():
        item_raw   = str(row.get("Item") or "")
        item_clean = _clean(item_raw)
        md_clean   = _clean(row.get(md_col)) if md_col else ""
        section    = (str(row.get(section_field) or "").strip()
                      if section_field in df.columns else "")
        section_clean = _clean(section)
        incumbent  = str(row.get(code_col) or "").strip() if code_col else ""
        if incumbent.lower() == "nan":
            incumbent = ""
        price    = sf._safe_float(row.get(price_field))
        qty      = sf._safe_float(row.get("Quantity")) if "Quantity" in df.columns else 0.0
        norm_price = (abs(price) / qty) if (price and qty and qty > 1) else abs(price or 0.0)
        is_zero  = _is_zero_priced(row, price)
        per_unit = _is_per_unit(row.get("Frequency"), item_raw)
        freq     = row.get("Frequency")

        # Cache the query/sims (many rows repeat verbatim — tiers, duplicates).
        skey = (item_clean, md_clean, section_clean)
        if skey not in _sim_cache:
            qv = _query_vec(item_clean, md_clean, section_clean)
            sims = _lex_sims(qv)
            _sim_cache[skey] = (sims, _best_sim_by_code(sims))
        sims, sim_by_code = _sim_cache[skey]

        # ── Candidate pool ──────────────────────────────────────────────────
        cands: dict = {}                 # code -> source tags
        def _add(code, tag):
            if code and code in by_code:   # MUST be a code billed in-scope
                cands.setdefault(code, set()).add(tag)

        # (a) exact in-scope price (the anchor) — match on |price| and |unit|.
        exact_codes: set = set()
        for amt in {_cents(price or 0.0), _cents(norm_price)}:
            b = price_hits.get(amt)
            if b:
                for rec in b["in"]:
                    _add(rec["code"], "exact-price")
                    exact_codes.add(rec["code"])
        # (b) ±25% buffer on net or unit price (from scope stats).
        if norm_price > 0:
            lo, hi = norm_price * (1 - NORM_PRICE_BUFFER), norm_price * (1 + NORM_PRICE_BUFFER)
            for code, st in by_code.items():
                for key in ("median_abs_net", "median_unit"):
                    m = st.get(key)
                    if m is not None and lo <= abs(m) <= hi:
                        _add(code, "buffer-price")
                        break
        # (c) description / keyword candidates (lexical over the scope corpus).
        for code, s in sorted(sim_by_code.items(), key=lambda kv: -kv[1])[:POOL_CAP]:
            if s >= 0.20:
                _add(code, "description")
        # (d) the incumbent — only if it's actually in scope.
        if incumbent:
            _add(incumbent, "incumbent")

        # Pre-rank the pool deterministically so the LLM sees the real contenders.
        def _prelim(code):
            st = by_code.get(code)
            s_lex = _clamp01(sim_by_code.get(code, 0.0))
            s_price = _price_signal(price if not is_zero else 0.0, st, is_zero,
                                    per_unit, code in exact_codes)
            return 0.6 * s_lex + 0.4 * s_price
        pool = sorted(cands.keys(), key=lambda c: -_prelim(c))[:POOL_CAP]

        rows_meta[idx] = {
            "item_raw": item_raw, "item_clean": item_clean, "md_clean": md_clean,
            "section": section, "incumbent": incumbent, "price": price,
            "norm_price": norm_price, "is_zero": is_zero, "per_unit": per_unit,
            "freq": freq, "sims": sims, "sim_by_code": sim_by_code,
            "pool": pool, "exact_codes": exact_codes, "cands_src": cands,
        }

        # LLM payload — top contenders only (bounds gateway cost / truncation).
        llm_pool = pool[:LLM_CAND_CAP]
        if llm_pool and (len(pool) > 1):
            gpt_payloads.append({
                "row_id": int(idx),
                "item": item_raw[:200],
                "matched_description": (str(row.get(md_col) or "")[:120] if md_col else ""),
                "section": section[:100],
                "contract_price": (0.0 if is_zero else (round(price, 2) if price else None)),
                "contract_frequency": _canon_freq(freq) or str(freq or "")[:30],
                "candidates": [{
                    "code": c,
                    "invoice_descriptions": agg["code_descs"].get(c, [])[:3],
                    "sample_net": (round(by_code[c]["median_net"], 2)
                                   if by_code.get(c, {}).get("median_net") is not None else None),
                    "n_lines": by_code.get(c, {}).get("n_lines", 0),
                    "product_hier": sorted(by_code.get(c, {}).get("prod_hiers", set()))[:2],
                } for c in llm_pool],
            })

    # 7) LLM semantic+context rerank (enhancement; never fatal).
    gpt = {}
    llm_ok = False
    if gpt_payloads:
        log(f"Semantic+context rerank of {len(gpt_payloads)} row(s) with {SEMANTIC_MODEL}…")
        try:
            client = make_client(api_key or VALIDATION_API_KEY)
            gpt = _llm_semantic_rank(client, gpt_payloads, log)
            llm_ok = any(v.get("ranked") for v in gpt.values())
        except Exception as e:
            log(f"  ⚠ Reranker unavailable ({e}); deterministic scoring only.")
            gpt = {}
    if not llm_ok:
        log("  Semantic scoring in lexical-only mode (LLM unavailable).")

    # 8) Per-row final scoring through the funnel (pool → top5 → top3 → final).
    results: dict = {}      # idx -> chosen rec dict
    for idx, meta in rows_meta.items():
        pool = meta["pool"]
        if not pool:
            results[idx] = None
            continue
        g = gpt.get(idx, {})
        llm_rel = {_norm_code(c): r for c, r in (g.get("ranked") or [])}
        sim_by_code = meta["sim_by_code"]
        price = meta["price"]; is_zero = meta["is_zero"]; per_unit = meta["per_unit"]
        exact_codes = meta["exact_codes"]

        scored = []
        for code in pool:
            st = by_code.get(code)
            s_lex = _clamp01(sim_by_code.get(code, 0.0))
            s_llm = llm_rel.get(_norm_code(code))
            # semantic+context: LLM primary, lexical floor; description anchor.
            if s_llm is not None:
                semctx = 0.7 * s_llm + 0.3 * s_lex
            else:
                semctx = s_lex
            if s_lex >= DESC_ANCHOR_SIM:
                semctx = max(semctx, s_lex)        # near-exact desc anchors
            is_exact = code in exact_codes
            s_price = _price_signal(price, st, is_zero, per_unit, is_exact)
            # Rule 6: an exact-price anchor only counts full when the code is
            # also plausible, OR the amount is highly specific (few codes share).
            if is_exact and semctx < PRICE_ANCHOR_MIN_SEM:
                specific = any(
                    price_hits.get(amt, {}).get("n_codes", 99) <= SPECIFIC_PRICE_MAX_CODES
                    for amt in {_cents(price or 0.0), _cents(meta["norm_price"])})
                if not specific:
                    s_price = min(s_price, 0.6)
            scored.append({"code": code, "semctx": _clamp01(semctx),
                           "price": _clamp01(s_price), "lex": s_lex,
                           "is_exact": is_exact, "stats": st})

        # Funnel: top5 by (sem+ctx, price) → add probability → top3 → frequency.
        scored.sort(key=lambda r: -(W_SEMCTX * r["semctx"] + W_PRICE * r["price"]))
        top5 = scored[:TOP5]
        prob = _probability(meta["sims"], [r["code"] for r in top5])
        for r in top5:
            r["prob"] = _clamp01(prob.get(r["code"], 0.0))
        top5.sort(key=lambda r: -(W_SEMCTX * r["semctx"] + W_PRICE * r["price"]
                                  + W_PROB * r["prob"]))
        top3 = top5[:TOP3]
        for r in top3:
            cadence = _infer_cadence(r["stats"])
            r["cadence"] = cadence
            r["freq"] = _clamp01(_freq_signal(meta["freq"], meta["item_raw"], cadence))
            r["final"] = (W_PRICE * r["price"] + W_SEMCTX * r["semctx"]
                          + W_PROB * r["prob"] + W_FREQ * r["freq"])
        top3.sort(key=lambda r: -r["final"])
        results[idx] = {"top3": top3, "all": scored}

    # 9) Tier consistency — same base fee + section → one code (best aggregate).
    groups: dict = {}
    for idx, meta in rows_meta.items():
        if results.get(idx) and results[idx]["top3"]:
            groups.setdefault(_base_fee_key(meta["item_raw"], meta["section"]), []).append(idx)
    tier_override: dict = {}
    for key, idxs in groups.items():
        if len(idxs) < 2:
            continue
        agg_score: dict = {}
        for idx in idxs:
            for r in results[idx]["top3"]:
                agg_score[r["code"]] = agg_score.get(r["code"], 0.0) + r["final"]
        if not agg_score:
            continue
        best_code = max(agg_score, key=agg_score.get)
        for idx in idxs:
            row_codes = {r["code"] for r in results[idx]["top3"]}
            if best_code in row_codes and best_code != results[idx]["top3"][0]["code"]:
                tier_override[idx] = best_code

    # 10) Assemble output.
    out = df.copy()
    for col in _NEW_COLUMNS:             # object dtype: holds str AND float
        out[col] = pd.Series([None] * len(out), index=out.index, dtype="object")

    band_counts = {"GREEN": 0, "YELLOW": 0, "RED": 0}
    n_changed = 0
    for idx, row in df.iterrows():
        meta = rows_meta.get(idx, {})
        incumbent = meta.get("incumbent", "")
        out.at[idx, "old_material_code"] = incumbent
        out.at[idx, "old_matched_description"] = (str(row.get(md_col) or "") if md_col else "")
        out.at[idx, "old_match_confidence"] = (row.get(conf_col) if conf_col else "")

        res = results.get(idx)
        if not res or not res["top3"]:
            # No in-scope candidate → keep incumbent, RED, never invent.
            out.at[idx, "new_material_code"] = incumbent
            out.at[idx, "new_matched_description"] = (str(row.get(md_col) or "") if md_col else "")
            out.at[idx, "new_validation_score"] = 0.0
            out.at[idx, "confidence_band"] = "RED"
            out.at[idx, "validation_route"] = "no-candidate"
            note = _no_candidate_reason(meta, price_hits, by_code)
            out.at[idx, "validation_reason"] = note
            band_counts["RED"] += 1
            continue

        top3 = res["top3"]
        chosen = top3[0]
        if idx in tier_override:
            chosen = next((r for r in top3 if r["code"] == tier_override[idx]), chosen)

        # Semantic-confidence gate: an unanchored, semantically-weak pick can't be
        # banded confident on the back of probability/frequency alone.
        score = chosen["final"]
        anchored = chosen["is_exact"] or chosen["lex"] >= DESC_ANCHOR_SIM
        if not anchored and chosen["semctx"] < SEM_CONF_FLOOR:
            score *= max(0.0, chosen["semctx"] / SEM_CONF_FLOOR)
        band = ("GREEN" if score >= GREEN_CUTOFF
                else "YELLOW" if score >= YELLOW_CUTOFF else "RED")
        band_counts[band] += 1

        fallback = next((r["code"] for r in top3 if r["code"] != chosen["code"]), "")
        # Description billed under the chosen code (real biller wording).
        stats = chosen.get("stats") or {}
        new_desc = (agg["code_descs"].get(chosen["code"], [None]) or [None])[0] \
            or stats.get("sample_desc", "")

        route = _route_label(chosen, idx in tier_override)
        g_reason = gpt.get(idx, {}).get("reason", "")
        reason = _compose_reason(chosen, incumbent, route, g_reason, llm_ok,
                                 meta, price_hits, band)

        if chosen["code"] != incumbent:
            n_changed += 1

        out.at[idx, "new_material_code"] = chosen["code"]
        out.at[idx, "new_matched_description"] = new_desc
        out.at[idx, "new_validation_score"] = round(score, 4)
        out.at[idx, "confidence_band"] = band
        out.at[idx, "fallback_material_code"] = fallback
        out.at[idx, "validation_reason"] = reason
        out.at[idx, "invoice_cadence"] = chosen.get("cadence", "Unknown")
        out.at[idx, "validation_route"] = route
        out.at[idx, "v_price"] = round(chosen["price"], 3)
        out.at[idx, "v_semantic_context"] = round(chosen["semctx"], 3)
        out.at[idx, "v_probability"] = round(chosen.get("prob", 0.0), 3)
        out.at[idx, "v_frequency"] = round(chosen.get("freq", 0.0), 3)
        out.at[idx, "top_candidates"] = ", ".join(
            f"{r['code']}:{round(r['final'], 2)}" for r in top3)

    # 11) Write + conditional formatting.
    out_path = output_path(client_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out.to_excel(str(out_path), index=False)
    _apply_formatting(out_path, list(out.columns))

    log(f"Wrote {out_path.name}: {len(out)} rows — "
        f"{band_counts['GREEN']} GREEN · {band_counts['YELLOW']} YELLOW · "
        f"{band_counts['RED']} RED · {n_changed} code(s) changed.")
    return {"status": "complete", "client": client_name, "rows": len(out),
            "output": str(out_path), "bands": band_counts}


# ── Reason / route helpers ────────────────────────────────────────────────────
def _route_label(chosen: dict, tiered: bool) -> str:
    if tiered:
        return "tier-consistent"
    if chosen["is_exact"] and chosen["semctx"] >= PRICE_ANCHOR_MIN_SEM:
        return "exact-price-anchor"
    if chosen["lex"] >= DESC_ANCHOR_SIM:
        return "description-anchor"
    if chosen["is_exact"]:
        return "exact-price+weak-semantic"
    if chosen["price"] >= 0.7 and chosen["semctx"] >= 0.4:
        return "price+semantic"
    return "semantic"


def _compose_reason(chosen, incumbent, route, llm_reason, llm_ok, meta,
                    price_hits, band="") -> str:
    parts = []
    if band == "RED":
        # Low-confidence pick: state that plainly and don't claim it "favours".
        parts.append("low-confidence: nearest in-scope billed code, but weak "
                     "evidence — review")
    elif llm_reason:
        parts.append(llm_reason)
    else:
        if chosen["semctx"] >= 0.75:
            parts.append("strong match to the client's invoice descriptions")
        elif chosen["semctx"] < 0.4:
            parts.append("weak semantic match")
        if chosen["is_exact"]:
            parts.append("billed at the exact contract amount")
        elif chosen["price"] >= 0.85:
            parts.append("price aligns with invoice history")
        if chosen.get("freq", 0) >= 0.85:
            parts.append("billing cadence matches the contract")
    if band != "RED" and chosen["code"] != incumbent:
        parts.append("invoice history favours this code over the original match")
    # Cross-scope diagnostic — exact amount exists ONLY outside this client.
    p = abs(meta.get("price") or 0.0)
    if p:
        b = price_hits.get(_cents(p))
        if b and not b["in"] and b["out"]:
            descs = ", ".join(sorted({r["desc"][:30] for r in b["out"]})[:2])
            parts.append(
                f"NOTE: ${p:,.2f} is billed only OUTSIDE this client's scope "
                f"({descs}) — verify sold-to/bill-to resolution")
    out = "; ".join(p for p in parts if p) or "scored from historical invoice usage"
    out = out[0].upper() + out[1:] if out else out
    if not llm_ok:
        out += " [lexical scoring - semantic model unavailable]"
    return out.rstrip(".") + f". [route: {route}]"


def _no_candidate_reason(meta, price_hits, by_code) -> str:
    inc = meta.get("incumbent", "")
    if inc and inc not in by_code:
        base = (f"Original code {inc} has no invoice history for this client's "
                "scope; no in-scope candidate found")
    else:
        base = "No invoice-supported candidate for this item in the client's scope"
    p = abs(meta.get("price") or 0.0)
    if p:
        b = price_hits.get(_cents(p))
        if b and not b["in"] and b["out"]:
            base += (f"; ${p:,.2f} is billed only outside this scope — verify "
                     "sold-to/bill-to resolution")
    return base + ". [route: no-candidate]"


def _apply_formatting(path: Path, columns: list) -> None:
    """Colour confidence_band + new_validation_score cells GREEN/YELLOW/RED."""
    try:
        wb = load_workbook(str(path))
        ws = wb.active
        try:
            band_i = columns.index("confidence_band") + 1
            score_i = columns.index("new_validation_score") + 1
        except ValueError:
            return
        fills = {"GREEN": _GREEN_FILL, "YELLOW": _YELLOW_FILL, "RED": _RED_FILL}
        for r in range(2, ws.max_row + 1):
            band = str(ws.cell(row=r, column=band_i).value or "").strip().upper()
            fill = fills.get(band)
            if fill:
                ws.cell(row=r, column=band_i).fill = fill
                ws.cell(row=r, column=score_i).fill = fill
        wb.save(str(path))
    except Exception:
        pass            # formatting is cosmetic — never fail the run over it


def _status_for(reason: str) -> str:
    r = reason.lower()
    if "matching" in r:
        return "no_matching"
    if "history" in r or "bill-to" in r or "sold-to" in r or "scope" in r:
        return "no_history"
    return "no_snowflake"
